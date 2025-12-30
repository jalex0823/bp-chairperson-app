from datetime import datetime, date, timedelta
from functools import wraps
import json
import os
import hashlib
import secrets
import base64
from datetime import timezone
from zoneinfo import ZoneInfo
import time

from flask import (
    Flask, render_template, redirect, url_for,
    request, flash, session, make_response, jsonify, abort, g, Response
)
from flask import send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from flask_wtf import FlaskForm
from flask_mail import Mail, Message
from wtforms import (
    StringField, TextAreaField, BooleanField,
    DateField, TimeField, PasswordField, SubmitField, IntegerField, RadioField, SelectField, FileField
)
from wtforms.validators import (
    DataRequired, Optional, Email, Length, NumberRange
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from icalendar import Calendar, Event, Alarm
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from dotenv import load_dotenv
import calendar
from io import BytesIO
import subprocess
import tempfile

# ReportLab depends on Pillow for some functionality; keep PDF/certificate features optional
# so auth/admin utilities can run in minimal environments.
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except Exception:
    letter = None
    canvas = None
    inch = None
    REPORTLAB_AVAILABLE = False

try:
    from reportlab.lib.utils import ImageReader
    REPORTLAB_IMAGE_READER_AVAILABLE = True
except Exception:
    ImageReader = None
    REPORTLAB_IMAGE_READER_AVAILABLE = False
from urllib.request import urlopen
from urllib.error import URLError, HTTPError

# Try to import Redis and Caching
try:
    from flask_caching import Cache
    CACHE_AVAILABLE = True
except ImportError:
    CACHE_AVAILABLE = False
    Cache = None

try:
    import redis
    REDIS_AVAILABLE = True
except ImportError:
    REDIS_AVAILABLE = False

# Load environment variables from .env file BEFORE importing Config so it can read envs
load_dotenv()

from config import Config

app = Flask(__name__)
app.config.from_object(Config)
db = SQLAlchemy(app)
mail = Mail(app)

# Configure caching
if CACHE_AVAILABLE:
    if REDIS_AVAILABLE and os.getenv('REDIS_URL'):
        # Use Redis for caching in production when REDIS_URL is set
        try:
            cache = Cache(app, config={
                'CACHE_TYPE': 'RedisCache',
                'CACHE_REDIS_URL': os.getenv('REDIS_URL'),
                'CACHE_DEFAULT_TIMEOUT': 300  # 5 minutes default
            })
            print("✓ Using Redis cache")
        except Exception as e:
            print(f"⚠ Redis connection failed, falling back to SimpleCache: {e}")
            cache = Cache(app, config={
                'CACHE_TYPE': 'SimpleCache',
                'CACHE_DEFAULT_TIMEOUT': 300
            })
    else:
        # Fallback to simple cache (no Redis available or configured)
        cache = Cache(app, config={
            'CACHE_TYPE': 'SimpleCache',
            'CACHE_DEFAULT_TIMEOUT': 300
        })
else:
    # No caching available - create a dummy cache object
    class DummyCache:
        def memoize(self, timeout=None):
            """Pass-through decorator when cache is not available"""
            def decorator(f):
                return f
            return decorator
        
        def cached(self, timeout=None):
            """Pass-through decorator when cache is not available"""
            def decorator(f):
                return f
            return decorator
        
        def clear(self):
            pass
    
    cache = DummyCache()
    print("⚠ Flask-Caching not available, caching disabled")

# Performance monitoring
@app.before_request
def before_request():
    g.start_time = time.time()

@app.after_request
def after_request(response):
    # Add performance headers
    if hasattr(g, 'start_time'):
        response_time = time.time() - g.start_time
        response.headers['X-Response-Time'] = f"{response_time:.3f}s"
    
    # Security headers
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    # Safari cookie compatibility headers
    # Ensure cookies are properly handled in Safari
    if 'Set-Cookie' in response.headers:
        # Safari requires explicit SameSite and Secure attributes
        # These are set in config, but we ensure the response respects them
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
        response.headers['Pragma'] = 'no-cache'
    
    return response

# Expose a simple asset version for cache-busting static resources
@app.context_processor
def inject_asset_version():
    ver = os.environ.get("ASSET_VERSION", "20251206")
    return {"asset_version": ver}

# Only start scheduler in development/local environment
# On Heroku, this runs in a separate worker process
if os.getenv('FLASK_ENV') != 'production':
    scheduler = BackgroundScheduler()
    scheduler.start()
else:
    scheduler = None

# Timezone helper functions
# Back Porch meetings are in US Eastern Time
EASTERN_TZ = ZoneInfo("America/New_York")

def get_eastern_now():
    """Get current datetime in Eastern Time."""
    return datetime.now(EASTERN_TZ)

def get_eastern_today():
    """Get current date in Eastern Time (not UTC)."""
    return get_eastern_now().date()

# External resources directory for chairing PDFs (configurable)
# Prefer project-local resources/pdfs so it works both locally and on Heroku.
DEFAULT_PDFS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources', 'pdfs')
EXTERNAL_PDFS_DIR = os.environ.get("EXTERNAL_PDFS_DIR", DEFAULT_PDFS_DIR)
SOURCE_MEETINGS_ICS_URL = os.environ.get("SOURCE_MEETINGS_ICS_URL")

# Profile image upload configuration
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5MB

def allowed_file(filename):
    """Check if file extension is allowed for profile images."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
SOURCE_MEETINGS_WEB_URL = os.environ.get("SOURCE_MEETINGS_WEB_URL")

# Optional: built-in static schedule (when no website calendar exists)
# Times are in local server time; title/description can be customized
STATIC_SCHEDULE_ENABLED = (os.environ.get("STATIC_SCHEDULE_ENABLED", "True").lower() == "true")
STATIC_SCHEDULE = {
    # Daily Literature-based Meeting at 17:30 (5:30 PM)
    "daily": {"enabled": True, "hour": 17, "minute": 30, "title": "Daily Literature-based Meeting", "description": "AA-approved literature only", "zoom_link": "Online"},
    # Women's: Saturday 08:30
    "women_sat": {"enabled": True, "weekday": 5, "hour": 8, "minute": 30, "title": "Women's Meeting", "description": "Women only", "zoom_link": "Online", "gender": "female"},
    # Co-ed: Sunday 08:30
    "coed_sun": {"enabled": True, "weekday": 6, "hour": 8, "minute": 30, "title": "Co-ed Meeting", "description": "Co-ed", "zoom_link": "Online"},
    # Men's: Sunday 15:30
    "men_sun": {"enabled": True, "weekday": 6, "hour": 15, "minute": 30, "title": "Men's Meeting", "description": "Men only", "zoom_link": "Online", "gender": "male"},
}

# Flask 3 removed before_first_request; database initialization is handled
# via Heroku release phase (see Procfile) and CLI command `flask --app app.py init-db`.

# ==========================
# MODELS
# ==========================

class User(db.Model):
    """
    Both admins and chairpersons.
    - is_admin = True: full access to manage meetings.
    - is_admin = False: regular chairperson account.
    """
    __tablename__ = "users"

    id = db.Column(db.Integer, primary_key=True)
    display_name = db.Column(db.String(80), nullable=False)  # e.g. "Jeff A."
    email = db.Column(db.String(255), unique=True, nullable=False, index=True)  # Index for login lookups
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False, index=True)  # Index for admin filtering
    sobriety_days = db.Column(db.Integer, nullable=True)
    agreed_guidelines = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)  # Index for user stats
    gender = db.Column(db.String(10), nullable=True, index=True)  # Index for gender filtering
    last_login = db.Column(db.DateTime, nullable=True, index=True)  # Index for activity tracking
    failed_login_attempts = db.Column(db.Integer, default=0)
    locked_until = db.Column(db.DateTime, nullable=True, index=True)  # Index for locked account queries
    profile_image = db.Column(db.Text, nullable=True)  # Store base64 encoded image data
    chair_points = db.Column(db.Integer, default=0, index=True)  # ChairPoints earned by chairing meetings
    password_reset_required = db.Column(db.Boolean, default=False)  # Force password change on next login

    chair_signups = db.relationship("ChairSignup", back_populates="user")
    availability_signups = db.relationship("ChairpersonAvailability", back_populates="user")

    @property
    def bp_id(self):
        """Generate Back Porch ID like BP-1001, BP-1002, etc."""
        return f"BP-{1000 + self.id}"

    def set_password(self, password: str):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)

    def has_role(self, role_name):
        """Check if user has a specific role."""
        return any(
            role.role == role_name and role.is_active and 
            (role.expires_at is None or role.expires_at > datetime.now(timezone.utc))
            for role in self.roles
        )

    def is_locked(self):
        """Check if user account is locked due to failed login attempts."""
        return self.locked_until and self.locked_until > datetime.utcnow()

    def lock_account(self, duration_minutes=30):
        """Lock user account for specified duration."""
        self.locked_until = datetime.utcnow() + timedelta(minutes=duration_minutes)
        db.session.commit()

    def unlock_account(self):
        """Unlock user account and reset failed attempts."""
        self.locked_until = None
        self.failed_login_attempts = 0
        db.session.commit()


class Meeting(db.Model):
    """
    A single Back Porch online meeting occurrence.
    Example: 'Back Porch Noon Meeting', 2025-12-08, 12:00 PM.
    """
    __tablename__ = "meetings"

    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=True)           # format / focus
    zoom_link = db.Column(db.String(500), nullable=True)      # meeting URL
    event_date = db.Column(db.Date, nullable=False, index=True)  # Index for date queries
    start_time = db.Column(db.Time, nullable=False, index=True)  # Index for time queries
    end_time = db.Column(db.Time, nullable=True)
    is_open = db.Column(db.Boolean, default=True, index=True)   # Index for filtering
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    gender_restriction = db.Column(db.String(10), nullable=True, index=True)  # Index for filtering
    meeting_type = db.Column(db.String(50), nullable=False, default='Regular', index=True)  # Index for filtering

    chair_signup = db.relationship(
        "ChairSignup",
        back_populates="meeting",
        uselist=False,
        cascade="all, delete-orphan"
    )

    @property
    def has_chair(self) -> bool:
        return self.chair_signup is not None

    @property
    def type_badge_class(self) -> str:
        """Return Bootstrap badge class for meeting type"""
        type_classes = {
            'Regular': 'bg-primary',
            'Special': 'bg-warning text-dark',
            'Holiday': 'bg-danger',
            'Workshop': 'bg-info'
        }
        return type_classes.get(self.meeting_type, 'bg-secondary')

    @property
    def type_icon(self) -> str:
        """Return Font Awesome icon class for meeting type"""
        type_icons = {
            'Regular': 'fas fa-users',
            'Special': 'fas fa-star',
            'Holiday': 'fas fa-holly-berry',
            'Workshop': 'fas fa-chalkboard-teacher'
        }
        return type_icons.get(self.meeting_type, 'fas fa-circle')


class ChairSignup(db.Model):
    """
    The record representing that a user has committed to chair a meeting.
    """
    __tablename__ = "chair_signups"

    id = db.Column(db.Integer, primary_key=True)
    meeting_id = db.Column(
        db.Integer,
        db.ForeignKey("meetings.id", ondelete="CASCADE"),
        unique=True,
        nullable=False
    )
    user_id = db.Column(
        db.Integer,
        db.ForeignKey("users.id", ondelete="CASCADE"),
        nullable=False
    )
    display_name_snapshot = db.Column(db.String(80), nullable=False)
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    meeting = db.relationship("Meeting", back_populates="chair_signup")
    user = db.relationship("User", back_populates="chair_signups")


class ChairpersonAvailability(db.Model):
    """
    Record when a user volunteers to chair on a specific date where no meeting exists yet.
    This allows users to express interest in chairing before meetings are scheduled.
    """
    __tablename__ = "chairperson_availability"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(
        db.Integer,
        db.ForeignKey("users.id", ondelete="CASCADE"),
        nullable=False
    )
    volunteer_date = db.Column(db.Date, nullable=False)
    time_preference = db.Column(db.String(50), nullable=True)  # "morning", "afternoon", "evening", "any"
    notes = db.Column(db.Text, nullable=True)
    display_name_snapshot = db.Column(db.String(80), nullable=False)
    is_active = db.Column(db.Boolean, default=True)  # Can be deactivated if converted to actual meeting
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", back_populates="availability_signups")

    __table_args__ = (db.UniqueConstraint('user_id', 'volunteer_date', name='uq_user_date_availability'),)


class AuditLog(db.Model):
    """
    Audit trail for important security events and administrative actions.
    """
    __tablename__ = "audit_logs"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True, index=True)  # Index for user filtering
    action = db.Column(db.String(100), nullable=False, index=True)  # Index for action filtering
    resource_type = db.Column(db.String(50), nullable=True, index=True)  # Index for resource filtering
    resource_id = db.Column(db.Integer, nullable=True, index=True)  # Index for resource lookups
    ip_address = db.Column(db.String(45), nullable=True, index=True)  # Index for IP tracking
    user_agent = db.Column(db.Text, nullable=True)
    details = db.Column(db.JSON, nullable=True)  # additional structured data
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), index=True)  # Index for time-based queries

    user = db.relationship("User", backref="audit_logs")


class SecurityToken(db.Model):
    """
    Security tokens for password resets, email verification, and two-factor authentication.
    """
    __tablename__ = "security_tokens"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    token_type = db.Column(db.String(20), nullable=False)  # password_reset, email_verify, totp_backup
    token_hash = db.Column(db.String(255), nullable=False)  # hashed token
    expires_at = db.Column(db.DateTime, nullable=False)
    used_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    user = db.relationship("User", backref="security_tokens")

    @classmethod
    def create_token(cls, user_id, token_type, expires_in_hours=24):
        """Create a new security token and return the unhashed version."""
        token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        
        security_token = cls(
            user_id=user_id,
            token_type=token_type,
            token_hash=token_hash,
            expires_at=datetime.now(timezone.utc) + timedelta(hours=expires_in_hours)
        )
        db.session.add(security_token)
        return token, security_token

    @classmethod
    def verify_token(cls, token, token_type):
        """Verify a token and mark it as used if valid."""
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        
        security_token = cls.query.filter_by(
            token_hash=token_hash,
            token_type=token_type,
            used_at=None
        ).filter(cls.expires_at > datetime.now(timezone.utc)).first()
        
        if security_token:
            security_token.used_at = datetime.now(timezone.utc)
            db.session.commit()
            return security_token.user
        return None


class UserRole(db.Model):
    """
    Role-based access control for enhanced security.
    """
    __tablename__ = "user_roles"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    role = db.Column(db.String(50), nullable=False)  # admin, moderator, chairperson, readonly
    granted_by = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    granted_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    expires_at = db.Column(db.DateTime, nullable=True)  # null for permanent roles
    is_active = db.Column(db.Boolean, default=True)

    user = db.relationship("User", foreign_keys=[user_id], backref="roles")
    granted_by_user = db.relationship("User", foreign_keys=[granted_by])


class BackupLog(db.Model):
    """
    Track database backup operations.
    """
    __tablename__ = "backup_logs"

    id = db.Column(db.Integer, primary_key=True)
    backup_type = db.Column(db.String(20), nullable=False)  # manual, scheduled, pre_update
    file_path = db.Column(db.String(500), nullable=True)
    file_size = db.Column(db.BigInteger, nullable=True)
    checksum = db.Column(db.String(64), nullable=True)  # SHA-256
    status = db.Column(db.String(20), nullable=False, default='started')  # started, completed, failed
    error_message = db.Column(db.Text, nullable=True)
    initiated_by = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    completed_at = db.Column(db.DateTime, nullable=True)

    initiated_by_user = db.relationship("User", backref="initiated_backups")


class QuizAttempt(db.Model):
    """
    Track user quiz attempts and scores for video training quizzes.
    """
    __tablename__ = "quiz_attempts"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    quiz_id = db.Column(db.String(50), nullable=False)  # 'registration_quiz' or 'hosting_quiz'
    score = db.Column(db.Integer, nullable=False)  # Percentage score (0-100)
    total_questions = db.Column(db.Integer, nullable=False)
    correct_answers = db.Column(db.Integer, nullable=False)
    passed = db.Column(db.Boolean, nullable=False)  # True if score >= 70%
    answers = db.Column(db.Text, nullable=True)  # JSON string of user answers
    completed_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    points_awarded = db.Column(db.Integer, default=0)  # ChairPoints earned

    user = db.relationship("User", backref="quiz_attempts")


# ==========================
# IMPORTERS / SYNC
# ==========================

def import_meetings_from_ics(ics_url: str, replace_future: bool = True) -> int:
    """Import meetings from an external iCal URL into our Meeting table.
    If replace_future is True, existing future meetings are removed before import.
    Returns number of meetings imported.
    """
    if not ics_url:
        raise ValueError("ICS URL not provided")
    # Support local filesystem paths in addition to URLs
    data = None
    if ics_url.lower().startswith('file://'):
        try:
            with urlopen(ics_url) as resp:
                data = resp.read()
        except (URLError, HTTPError) as e:
            raise RuntimeError(f"Failed to fetch ICS (file URL): {e}")
    elif os.path.exists(ics_url):
        try:
            with open(ics_url, 'rb') as f:
                data = f.read()
        except Exception as e:
            raise RuntimeError(f"Failed to read local ICS file: {e}")
    else:
        try:
            with urlopen(ics_url) as resp:
                data = resp.read()
        except (URLError, HTTPError) as e:
            raise RuntimeError(f"Failed to fetch ICS: {e}")

    try:
        cal = Calendar.from_ical(data)
    except Exception as e:
        raise RuntimeError(f"Invalid ICS content: {e}")

    imported = 0
    if replace_future:
        db.session.query(Meeting).filter(Meeting.event_date >= date.today()).delete()
        db.session.commit()

    for component in cal.walk():
        if component.name != 'VEVENT':
            continue

        # Dates
        dtstart = component.get('dtstart')
        dtend = component.get('dtend')
        if not dtstart:
            continue
        start_val = dtstart.dt
        end_val = dtend.dt if dtend else None
        # Normalize to date + time
        if hasattr(start_val, 'date') and hasattr(start_val, 'time'):
            event_date_val = start_val.date()
            start_time_val = start_val.time()
        else:
            # start_val may be a date
            event_date_val = start_val
            start_time_val = None
        end_time_val = None
        if end_val:
            if hasattr(end_val, 'time'):
                end_time_val = end_val.time()

        title = str(component.get('summary') or 'Back Porch Meeting')
        location = str(component.get('location') or '')
        description = str(component.get('description') or '')

        m = Meeting(
            title=title,
            description=description or None,
            zoom_link=location or None,
            event_date=event_date_val,
            start_time=start_time_val or (datetime.min + timedelta(hours=12)).time(),
            end_time=end_time_val,
            is_open=True,
            gender_restriction=None,
        )
        db.session.add(m)
        imported += 1

    db.session.commit()
    return imported


def seed_meetings_from_static_schedule(weeks: int = 12, replace_future: bool = True) -> int:
    """Generate meetings for the next N weeks based on STATIC_SCHEDULE.
    If replace_future is True, clear future meetings first.
    Returns number of meetings created.
    """
    if not STATIC_SCHEDULE_ENABLED:
        return 0
    if replace_future:
        db.session.query(Meeting).filter(Meeting.event_date >= date.today()).delete()
        db.session.commit()

    created = 0
    today = date.today()

    # Daily
    daily = STATIC_SCHEDULE.get("daily", {})
    if daily.get("enabled"):
        for i in range(weeks * 7):
            d = today + timedelta(days=i)
            start_t = datetime.min.replace(hour=daily["hour"], minute=daily["minute"]).time()
            m = Meeting(
                title=daily["title"],
                description=daily.get("description"),
                zoom_link=daily.get("zoom_link"),
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=daily["hour"], minute=daily["minute"]) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction=None,
            )
            db.session.add(m)
            created += 1

    def iter_weekdays(start: date, weekday: int, count: int):
        days_ahead = (weekday - start.weekday()) % 7
        d = start + timedelta(days=days_ahead)
        for _ in range(count):
            yield d
            d += timedelta(days=7)

    # Saturday women's
    w = STATIC_SCHEDULE.get("women_sat", {})
    if w.get("enabled"):
        for d in iter_weekdays(today, w["weekday"], weeks):
            start_t = datetime.min.replace(hour=w["hour"], minute=w["minute"]).time()
            m = Meeting(
                title=w["title"],
                description=w.get("description"),
                zoom_link=w.get("zoom_link"),
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=w["hour"], minute=w["minute"]) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction=w.get("gender"),
            )
            db.session.add(m)
            created += 1

    # Sunday co-ed
    c = STATIC_SCHEDULE.get("coed_sun", {})
    if c.get("enabled"):
        for d in iter_weekdays(today, c["weekday"], weeks):
            start_t = datetime.min.replace(hour=c["hour"], minute=c["minute"]).time()
            m = Meeting(
                title=c["title"],
                description=c.get("description"),
                zoom_link=c.get("zoom_link"),
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=c["hour"], minute=c["minute"]) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction=None,
            )
            db.session.add(m)
            created += 1

    # Sunday men's
    mconf = STATIC_SCHEDULE.get("men_sun", {})
    if mconf.get("enabled"):
        for d in iter_weekdays(today, mconf["weekday"], weeks):
            start_t = datetime.min.replace(hour=mconf["hour"], minute=mconf["minute"]).time()
            m = Meeting(
                title=mconf["title"],
                description=mconf.get("description"),
                zoom_link=mconf.get("zoom_link"),
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=mconf["hour"], minute=mconf["minute"]) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction=mconf.get("gender"),
            )
            db.session.add(m)
            created += 1

    db.session.commit()
    return created


def import_meetings_from_webpage(page_url: str, weeks: int = 12, replace_future: bool = True) -> int:
    """Scrape meeting schedule from an external HTML page and generate meetings for the next N weeks.
    This is resilient to minor content changes by looking for known phrases and time patterns.
    The current site schedule includes:
      - Daily Literature-based Meeting at 5:30 PM MT (co-ed)
      - Saturday Women's at 8:30 AM MT
      - Sunday Co-ed at 8:30 AM MT
      - Sunday Men's at 3:30 PM MT
    If parsing fails, no meetings are created.
    """
    if not page_url:
        raise ValueError("Web page URL not provided")

    # Fetch page
    try:
        with urlopen(page_url) as resp:
            html = resp.read().decode('utf-8', errors='ignore')
    except (URLError, HTTPError) as e:
        raise RuntimeError(f"Failed to fetch web page: {e}")

    import re
    text = re.sub(r"<[^>]+>", " ", html)  # strip tags to plain text
    text = re.sub(r"\s+", " ", text).lower()

    # Helper to detect presence of phrases
    def has(*phrases):
        return all(p.lower() in text for p in phrases)

    # Try to extract times with regex; default to known if phrases exist
    def find_time(patterns, default_hm):
        for pat in patterns:
            m = re.search(pat, text)
            if m:
                # Normalize hour/minute and am/pm
                hh = int(m.group('hour'))
                mm = int(m.group('min')) if m.group('min') else 0
                ampm = (m.group('ampm') or '').lower()
                if ampm == 'pm' and hh != 12:
                    hh += 12
                if ampm == 'am' and hh == 12:
                    hh = 0
                return hh, mm
        return default_hm

    # Patterns like "5:30 PM" or "5 PM"
    time_patterns = [r"(?P<hour>\d{1,2})(:(?P<min>\d{2}))?\s*(?P<ampm>am|pm)\s*mt"]

    # Determine configured schedule from page content (fallbacks to defaults)
    daily_enabled = has('daily') and has('literature')
    daily_hm = find_time(time_patterns, (17, 30))

    women_enabled = has("women") and has("saturday")
    women_hm = find_time(time_patterns, (8, 30))

    coed_enabled = has("co-ed") and has("sunday")
    coed_hm = find_time(time_patterns, (8, 30))

    men_enabled = has("men") and has("sunday")
    men_hm = find_time(time_patterns, (15, 30))

    if replace_future:
        db.session.query(Meeting).filter(Meeting.event_date >= date.today()).delete()
        db.session.commit()

    created = 0
    today = date.today()

    # Daily
    if daily_enabled:
        hh, mm = daily_hm
        for i in range(weeks * 7):
            d = today + timedelta(days=i)
            start_t = datetime.min.replace(hour=hh, minute=mm).time()
            m = Meeting(
                title="Daily Literature-based Meeting",
                description="AA-approved literature only",
                zoom_link="Online",
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=hh, minute=mm) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction=None,
            )
            db.session.add(m)
            created += 1

    def iter_weekdays(start: date, weekday: int, count: int):
        days_ahead = (weekday - start.weekday()) % 7
        d = start + timedelta(days=days_ahead)
        for _ in range(count):
            yield d
            d += timedelta(days=7)

    # Saturday women's (weekday=5)
    if women_enabled:
        hh, mm = women_hm
        for d in iter_weekdays(today, 5, weeks):
            start_t = datetime.min.replace(hour=hh, minute=mm).time()
            m = Meeting(
                title="Women's Meeting",
                description="Women only",
                zoom_link="Online",
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=hh, minute=mm) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction='female',
            )
            db.session.add(m)
            created += 1

    # Sunday co-ed (weekday=6)
    if coed_enabled:
        hh, mm = coed_hm
        for d in iter_weekdays(today, 6, weeks):
            start_t = datetime.min.replace(hour=hh, minute=mm).time()
            m = Meeting(
                title="Co-ed Meeting",
                description="Co-ed",
                zoom_link="Online",
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=hh, minute=mm) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction=None,
            )
            db.session.add(m)
            created += 1

    # Sunday men's (weekday=6)
    if men_enabled:
        hh, mm = men_hm
        for d in iter_weekdays(today, 6, weeks):
            start_t = datetime.min.replace(hour=hh, minute=mm).time()
            m = Meeting(
                title="Men's Meeting",
                description="Men only",
                zoom_link="Online",
                event_date=d,
                start_time=start_t,
                end_time=(datetime.min.replace(hour=hh, minute=mm) + timedelta(hours=1)).time(),
                is_open=True,
                gender_restriction='male',
            )
            db.session.add(m)
            created += 1

    db.session.commit()
    return created


# ==========================
# FORMS
# ==========================

class RegisterForm(FlaskForm):
    access_code = StringField("Access Code", validators=[Optional(), Length(max=64)])
    display_name = StringField(
        "Name to display (first name & last initial or alias)",
        validators=[DataRequired(), Length(max=80)]
    )
    email = StringField("Email", validators=[DataRequired(), Email(), Length(max=255)])
    password = PasswordField(
        "Password",
        validators=[DataRequired(), Length(min=6, message="At least 6 characters")]
    )
    # Change from approx. days to sobriety date; we'll auto-calc days server-side
    sobriety_date = DateField(
        "Sobriety Date (optional)",
        validators=[Optional()],
        format="%Y-%m-%d"
    )
    agreed_guidelines = BooleanField(
        "I have at least ~90 days sober and am working with a sponsor (suggested), and I have read the chairperson guidelines.",
        validators=[DataRequired(message="Please confirm you have read and agree to the guidelines.")]
    )
    # Gender selection for meeting restrictions
    gender = RadioField(
        "Gender",
        choices=[('male','Male'),('female','Female')],
        validators=[DataRequired(message="Please select your gender for meeting eligibility.")]
    )
    submit = SubmitField("Create Chairperson Account")

class AccessCodeFormMixin:
    from wtforms import StringField
    access_code = StringField("Access Code", validators=[Optional(), Length(max=64)])


class LoginForm(FlaskForm):
    email = StringField("Email", validators=[DataRequired(), Email()])
    password = PasswordField("Password", validators=[DataRequired()])
    submit = SubmitField("Log In")


class MeetingForm(FlaskForm):
    title = StringField("Meeting Name", validators=[DataRequired(), Length(max=255)])
    description = TextAreaField("Description / Format", validators=[Optional()])
    zoom_link = StringField("Online Meeting Link", validators=[Optional(), Length(max=500)])
    event_date = DateField("Date", validators=[DataRequired()], format="%Y-%m-%d")
    start_time = TimeField("Start Time", validators=[DataRequired()], format="%H:%M")
    end_time = TimeField("End Time (optional)", validators=[Optional()], format="%H:%M")
    is_open = BooleanField("Open for chair sign-ups?", default=True)
    gender_restriction = SelectField(
        "Gender Restriction",
        choices=[('','None'),('male','Men only'),('female','Women only')],
        validators=[Optional()],
        default=''
    )
    meeting_type = SelectField(
        "Meeting Type",
        choices=[
            ('Regular', 'Regular Meeting'),
            ('Special', 'Special Meeting'),
            ('Holiday', 'Holiday Meeting'),
            ('Workshop', 'Workshop/Study')
        ],
        validators=[DataRequired()],
        default='Regular'
    )
    submit = SubmitField("Save Meeting")


class ChairSignupForm(FlaskForm):
    notes = TextAreaField(
        "Notes (optional)",
        validators=[Optional(), Length(max=500)]
    )
    submit = SubmitField("Sign Up to Chair This Meeting")


class ChairpersonAvailabilityForm(FlaskForm):
    time_preference = RadioField(
        "Time Preference",
        choices=[
            ('any', 'Any time'),
            ('morning', 'Morning (before 12 PM)'),
            ('afternoon', 'Afternoon (12 PM - 6 PM)'),
            ('evening', 'Evening (after 6 PM)')
        ],
        default='any',
        validators=[DataRequired()]
    )
    notes = TextAreaField(
        "Notes (optional)",
        validators=[Optional(), Length(max=500)],
        description="Any additional notes about your availability or preferences"
    )
    submit = SubmitField("Volunteer for This Date")
    submit = SubmitField("Volunteer for This Date")


# ==========================
# SECURITY FUNCTIONS
# ==========================

def log_audit_event(action, user_id=None, resource_type=None, resource_id=None, details=None):
    """Log security and administrative events."""
    try:
        audit_log = AuditLog(
            user_id=user_id,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            ip_address=request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr),
            user_agent=request.environ.get('HTTP_USER_AGENT'),
            details=details
        )
        db.session.add(audit_log)
        db.session.commit()
    except Exception as e:
        print(f"Failed to log audit event: {e}")


def require_role(role_name):
    """Decorator to require specific role for access."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = get_current_user()
            if not user:
                flash("Please log in to access this page.", "warning")
                return redirect(url_for('login'))
            
            if not user.has_role(role_name) and not user.is_admin:
                log_audit_event('unauthorized_access_attempt', user.id, details={'required_role': role_name, 'requested_url': request.url})
                flash("You don't have permission to access this page.", "danger")
                abort(403)
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def require_login(f):
    """Decorator to require user to be logged in."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = get_current_user()
        if not user:
            flash("Please log in to access this page.", "warning")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def backup_database(backup_type='manual', initiated_by_user_id=None):
    """Create a database backup and log the operation."""
    backup_log = BackupLog(
        backup_type=backup_type,
        initiated_by=initiated_by_user_id,
        status='started'
    )
    db.session.add(backup_log)
    db.session.commit()
    
    try:
        # Create backup filename with timestamp
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        backup_filename = f"backup_{backup_type}_{timestamp}.sql"
        backup_path = os.path.join(tempfile.gettempdir(), backup_filename)
        
        # Get database URL from config
        db_url = app.config['SQLALCHEMY_DATABASE_URI']
        
        if db_url.startswith('sqlite:///'):
            # SQLite backup
            import shutil
            source_path = db_url.replace('sqlite:///', '')
            shutil.copy2(source_path, backup_path.replace('.sql', '.db'))
            backup_path = backup_path.replace('.sql', '.db')
        else:
            # MySQL/PostgreSQL backup using mysqldump/pg_dump
            if 'mysql' in db_url:
                # Parse MySQL connection details
                import urllib.parse as urlparse
                parsed = urlparse.urlparse(db_url)
                username = parsed.username
                password = parsed.password
                hostname = parsed.hostname
                port = parsed.port or 3306
                database = parsed.path[1:]  # Remove leading /
                
                cmd = [
                    'mysqldump',
                    f'--host={hostname}',
                    f'--port={port}',
                    f'--user={username}',
                    f'--password={password}',
                    database
                ]
                
                with open(backup_path, 'w') as f:
                    subprocess.run(cmd, stdout=f, check=True)
            else:
                raise ValueError(f"Unsupported database type in URL: {db_url}")
        
        # Calculate file size and checksum
        file_size = os.path.getsize(backup_path)
        with open(backup_path, 'rb') as f:
            checksum = hashlib.sha256(f.read()).hexdigest()
        
        # Update backup log
        backup_log.file_path = backup_path
        backup_log.file_size = file_size
        backup_log.checksum = checksum
        backup_log.status = 'completed'
        backup_log.completed_at = datetime.now(timezone.utc)
        db.session.commit()
        
        log_audit_event('database_backup', initiated_by_user_id, details={
            'backup_type': backup_type,
            'file_size': file_size,
            'checksum': checksum
        })
        
        return backup_log
        
    except Exception as e:
        backup_log.status = 'failed'
        backup_log.error_message = str(e)
        backup_log.completed_at = datetime.now(timezone.utc)
        db.session.commit()
        
        log_audit_event('database_backup_failed', initiated_by_user_id, details={
            'backup_type': backup_type,
            'error': str(e)
        })
        
        raise


def check_rate_limit(action, user_id, limit_per_hour=10):
    """Simple rate limiting for sensitive actions."""
    cutoff = datetime.now(timezone.utc) - timedelta(hours=1)
    recent_attempts = AuditLog.query.filter(
        AuditLog.user_id == user_id,
        AuditLog.action == action,
        AuditLog.created_at > cutoff
    ).count()
    
    return recent_attempts < limit_per_hour


# ==========================
# AUTH HELPERS
# ==========================

def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    try:
        return User.query.get(user_id)
    except Exception:
        # If database is unavailable, user is not logged in
        return None


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            flash("Please log in to continue.", "warning")
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapper


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user or not user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("index"))
        return view_func(*args, **kwargs)
    return wrapper


# ==========================
# PUBLIC ROUTES
# ==========================

@app.context_processor
def inject_globals():
    try:
        return {"current_user": get_current_user()}
    except Exception:
        # If database is unavailable, provide None user
        return {"current_user": None}

# Ensure there are always meetings to show: if database is empty and
# static schedule is enabled, auto-seed a minimal horizon on the first request.
# Flask 3.x removed before_first_request; emulate a one-time hook using a sentinel.
_auto_seed_done = False

@app.before_request
def ensure_meetings_exist():
    global _auto_seed_done
    if _auto_seed_done:
        return
    
    # Skip database operations for API endpoints that don't need them
    if request.endpoint and request.endpoint.startswith('api_'):
        return
    
    # Skip for specific endpoints that should work without database
    skip_endpoints = ['api_validate_registration_key']
    if request.endpoint in skip_endpoints:
        return
        
    try:
        total = Meeting.query.count()
        if total == 0 and app.config.get('STATIC_SCHEDULE_ENABLED', True):
            # Seed next 12 weeks so the homepage and calendar never look empty
            seed_meetings_from_static_schedule(weeks=12, replace_future=True)
            app.logger.info("Auto-seeded 12 weeks from static schedule (DB was empty).")
        _auto_seed_done = True
    except Exception as e:
        app.logger.warning(f"Auto-seed skipped due to error: {e}")
        _auto_seed_done = True

@app.route("/favicon.ico")
def favicon():
    """Serve favicon from static folder."""
    return send_from_directory(
        os.path.join(app.root_path, 'static', 'img'),
        'favicon.ico',
        mimetype='image/vnd.microsoft.icon'
    )

@app.route("/")
def index():
    """List upcoming meetings grouped by date for calendar view."""
    today = get_eastern_today()
    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= today)
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )
    # Group meetings by date
    meetings_by_date = {}
    for m in meetings:
        date_str = m.event_date.strftime('%Y-%m-%d')
        if date_str not in meetings_by_date:
            meetings_by_date[date_str] = []
        meetings_by_date[date_str].append(m)
    return render_template("index.html", meetings_by_date=meetings_by_date)


@app.route("/chair-resources")
def chair_resources():
    """Chairperson resources page including external PDF references."""
    pdfs = []
    try:
        if os.path.isdir(EXTERNAL_PDFS_DIR):
            for name in sorted(os.listdir(EXTERNAL_PDFS_DIR)):
                if name.lower().endswith('.pdf') and not name.startswith('.'):
                    pdfs.append({
                        'name': name,
                        'url': url_for('serve_pdf', filename=name)
                    })
    except Exception:
        pdfs = []
    return render_template("chair_resources.html", pdfs=pdfs)

@app.route("/host-signup-instructions")
def host_signup_instructions():
    """Host Sign-Up Instructions page with clear steps for hosting/zoom duties."""
    return render_template("host_signup_instructions.html")


@app.route("/registration-instructions")
def registration_instructions():
    """Registration Instructions page explaining how to register and what to expect."""
    return render_template("registration_instructions.html")


@app.route('/resources/pdfs/<path:filename>')
def serve_pdf(filename):
    # Only allow .pdf files from the configured directory
    if not filename.lower().endswith('.pdf'):
        return "Not allowed", 403
    if not os.path.isdir(EXTERNAL_PDFS_DIR):
        return "Resources folder not found", 404
    return send_from_directory(EXTERNAL_PDFS_DIR, filename, as_attachment=False)


@app.route("/meetings/today")
def meetings_today():
    """Public page listing today's meetings with time, title, and chair name."""
    today = get_eastern_today()
    meetings = (
        Meeting.query
        .filter(Meeting.event_date == today)
        .order_by(Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )
    return render_template("today_meetings.html", meetings=meetings, today=today)


@app.route("/calendar")
def calendar_view():
    """Monthly calendar view showing meeting title, host, and time per day."""
    # Determine month to show from query params, default to current month
    today = get_eastern_today()
    year = int(request.args.get("year", today.year))
    month = int(request.args.get("month", today.month))
    q = (request.args.get("q", "") or "").strip().lower()

    # Get all meetings for the month
    start_date = date(year, month, 1)
    # Compute last day of month
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= start_date, Meeting.event_date <= end_date)
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )

    # Apply simple search filter (title or chair name)
    if q:
        def _matches(m):
            title_match = q in (m.title or "").lower()
            chair_name = (
                m.chair_signup.display_name_snapshot if m.chair_signup else ""
            )
            chair_match = q in chair_name.lower()
            return title_match or chair_match
        meetings = [m for m in meetings if _matches(m)]

    # Map meetings by date
    meetings_by_date = {}
    for m in meetings:
        meetings_by_date.setdefault(m.event_date, []).append(m)

    # Build weeks for the month
    cal = calendar.Calendar(firstweekday=6)  # Sunday start
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        week_cells = []
        for d in week:
            in_month = d.month == month
            week_cells.append({
                "date": d if in_month else None,  # hide out-of-month day number
                "in_month": in_month,
                "meetings": meetings_by_date.get(d, []) if in_month else [],
            })
        weeks.append(week_cells)

    # Previous/next month links
    prev_month = month - 1
    prev_year = year
    if prev_month == 0:
        prev_month = 12
        prev_year -= 1
    next_month = month + 1
    next_year = year
    if next_month == 13:
        next_month = 1
        next_year += 1

    return render_template(
        "calendar.html",
        year=year,
        month=month,
        month_name=calendar.month_name[month],
        weeks=weeks,
        prev_year=prev_year,
        prev_month=prev_month,
        next_year=next_year,
        next_month=next_month,
        today=today,
        q=(request.args.get("q", "") or ""),
        allow_nav=bool(get_current_user()),
    )


@app.route("/calendar/display")
def calendar_display():
    """Read-only display calendar: always shows current month, no navigation."""
    today = get_eastern_today()
    year = today.year
    month = today.month

    # Build month data (reuse logic from calendar_view)
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= start_date, Meeting.event_date <= end_date)
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )

    meetings_by_date = {}
    for m in meetings:
        meetings_by_date.setdefault(m.event_date, []).append(m)

    cal = calendar.Calendar(firstweekday=6)
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        week_cells = []
        for d in week:
            in_month = d.month == month
            week_cells.append({
                "date": d if in_month else None,
                "in_month": in_month,
                "meetings": meetings_by_date.get(d, []) if in_month else [],
            })
        weeks.append(week_cells)

    # prev/next computed but not shown
    prev_month = month - 1
    prev_year = year if prev_month != 0 else year - 1
    if prev_month == 0:
        prev_month = 12
    next_month = month + 1
    next_year = year if next_month != 13 else year + 1
    if next_month == 13:
        next_month = 1

    return render_template(
        "calendar.html",
        year=year,
        month=month,
        month_name=calendar.month_name[month],
        weeks=weeks,
        prev_year=prev_year,
        prev_month=prev_month,
        next_year=next_year,
        next_month=next_month,
        today=today,
        q="",
        allow_nav=False,
    )


@app.route("/calendar/ics")
def calendar_month_ics():
    """Export meetings for a specific month as iCal (.ics) file."""
    today = get_eastern_today()
    year = int(request.args.get("year", today.year))
    month = int(request.args.get("month", today.month))

    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= start_date, Meeting.event_date <= end_date)
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )

    cal = Calendar()
    cal.add('prodid', '-//Back Porch Meetings//backporchmeetings.org//')
    cal.add('version', '2.0')
    cal.add('x-wr-calname', f'Back Porch Chairperson Calendar - {calendar.month_name[month]} {year}')

    for m in meetings:
        event = Event()
        start_dt = datetime.combine(m.event_date, m.start_time)
        end_dt = datetime.combine(m.event_date, m.end_time) if m.end_time else start_dt + timedelta(hours=1)
        event.add('summary', m.title)
        event.add('dtstart', start_dt)
        event.add('dtend', end_dt)
        event.add('location', m.zoom_link or 'Online')
        event.add('description', f"{m.description or ''}\n\nChair: {m.chair_signup.display_name_snapshot + ' (' + m.chair_signup.user.bp_id + ')' if m.chair_signup else 'No chair yet'}")
        event.add('url', url_for('meeting_detail', meeting_id=m.id, _external=True))
        event.add('uid', f"meeting-{m.id}@backporchmeetings.org")
        cal.add_component(event)

    response = make_response(cal.to_ical())
    response.headers['Content-Type'] = 'text/calendar; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=backporch-calendar-{year}-{month}.ics'
    return response


@app.route("/calendar/day-ics")
def calendar_day_ics():
    """Export meetings for a specific day as iCal (.ics) file."""
    today = get_eastern_today()
    date_str = request.args.get("date")
    
    if date_str:
        try:
            event_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            event_date = today
    else:
        event_date = today

    meetings = (
        Meeting.query
        .filter(Meeting.event_date == event_date)
        .order_by(Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )

    cal = Calendar()
    cal.add('prodid', '-//Back Porch Meetings//backporchmeetings.org//')
    cal.add('version', '2.0')
    cal.add('x-wr-calname', f'Back Porch Meetings - {event_date.strftime("%B %d, %Y")}')

    for m in meetings:
        event = Event()
        start_dt = datetime.combine(m.event_date, m.start_time)
        end_dt = datetime.combine(m.event_date, m.end_time) if m.end_time else start_dt + timedelta(hours=1)
        event.add('summary', m.title)
        event.add('dtstart', start_dt)
        event.add('dtend', end_dt)
        event.add('location', m.zoom_link or 'Online')
        event.add('description', f"{m.description or ''}\n\nChair: {m.chair_signup.display_name_snapshot + ' (' + m.chair_signup.user.bp_id + ')' if m.chair_signup else 'No chair yet'}")
        event.add('url', url_for('meeting_detail', meeting_id=m.id, _external=True))
        event.add('uid', f"meeting-{m.id}@backporchmeetings.org")
        cal.add_component(event)

    response = make_response(cal.to_ical())
    response.headers['Content-Type'] = 'text/calendar; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=backporch-{event_date.strftime("%Y-%m-%d")}.ics'
    return response


@app.route("/meeting/<int:meeting_id>", methods=["GET", "POST"])
def meeting_detail(meeting_id):
    """Show a single meeting and allow chair sign-up if open & unclaimed."""
    user = None
    meeting = None
    form = None
    
    try:
        meeting = Meeting.query.options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user)).filter_by(id=meeting_id).first()
        if not meeting:
            flash("Meeting not found.", "danger")
            return redirect(url_for("calendar_view"))
        
        user = get_current_user()
        form = ChairSignupForm()
        
        app.logger.info(f"Loading meeting {meeting_id}: {meeting.title}, has_chair={meeting.has_chair}, user={'logged_in' if user else 'anonymous'}")
        
    except Exception as e:
        app.logger.error(f"Error loading meeting {meeting_id}: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        flash("Error loading meeting details. Please try again.", "danger")
        return redirect(url_for("calendar_view"))

    if request.method == "POST" and form.validate_on_submit():
        if not user:
            flash("Please log in or create a chairperson account first.", "warning")
            return redirect(url_for("login", next=request.path))

        if not meeting.is_open:
            flash("This meeting is not currently open for chair sign-ups.", "danger")
        elif meeting.has_chair:
            flash("This meeting already has a chair. Thank you for your willingness!", "danger")
        else:
            signup = ChairSignup(
                meeting_id=meeting.id,
                user_id=user.id,
                display_name_snapshot=user.display_name,
                notes=form.notes.data.strip() if form.notes.data else None,
            )
            db.session.add(signup)
            db.session.commit()
            
            # Send confirmation email immediately
            try:
                send_chair_confirmation(signup)
                print(f"Sent confirmation email to {user.email} for meeting {meeting.id}")
            except Exception as e:
                print(f"Failed to send confirmation email: {e}")
            
            # Schedule reminder email 24 hours before meeting (only in development)
            if scheduler:
                reminder_time = datetime.combine(meeting.event_date, meeting.start_time) - timedelta(hours=24)
                if reminder_time > datetime.now():
                    scheduler.add_job(
                        send_chair_reminder,
                        'date',
                        run_date=reminder_time,
                        args=[meeting.id],
                        id=f"reminder-{meeting.id}",
                        replace_existing=True
                    )
            
            flash("Thank you for chairing this Back Porch meeting!", "success")
            return redirect(url_for("meeting_detail", meeting_id=meeting.id))

    try:
        app.logger.info(f"Rendering meeting_detail template for meeting {meeting_id}")
        return render_template("meeting_detail.html", meeting=meeting, form=form, user=user)
    except Exception as e:
        app.logger.error(f"Error rendering meeting_detail template for meeting {meeting_id}: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        # Don't show error to user since signup was successful - just redirect
        return redirect(url_for("dashboard"))


@app.route("/meeting/<int:meeting_id>/cancel", methods=["POST"])
@login_required
def cancel_chair_signup(meeting_id):
    """Allow user to cancel their chair signup for a meeting."""
    from app import ChairSignup, cache
    
    user = get_current_user()
    meeting = Meeting.query.get_or_404(meeting_id)
    
    # Find the user's chair signup for this meeting
    signup = ChairSignup.query.filter_by(
        meeting_id=meeting_id,
        user_id=user.id
    ).first()
    
    if not signup:
        flash("You don't have a chair signup for this meeting.", "warning")
        return redirect(url_for("meeting_detail", meeting_id=meeting_id))
    
    try:
        db.session.delete(signup)
        db.session.commit()
        
        # Clear cache to ensure dashboard and calendar reflect the change
        try:
            cache.clear()
        except Exception as cache_error:
            app.logger.warning(f"Failed to clear cache: {cache_error}")
        
        flash("You have successfully withdrawn from chairing this meeting.", "success")
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error canceling signup: {e}")
        flash(f"Error canceling signup: {str(e)}", "danger")
    
    return redirect(url_for("dashboard"))


@app.route("/calendar.ics")
def calendar_ics():
    """Export meetings as iCal calendar feed."""
    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= date.today())
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )
    
    cal = Calendar()
    cal.add('prodid', '-//Back Porch Meetings//backporchmeetings.org//')
    cal.add('version', '2.0')
    cal.add('x-wr-calname', 'Back Porch Chairperson Calendar')
    
    for m in meetings:
        event = Event()
        start_dt = datetime.combine(m.event_date, m.start_time)
        end_dt = datetime.combine(m.event_date, m.end_time) if m.end_time else start_dt + timedelta(hours=1)
        
        event.add('summary', m.title)
        event.add('dtstart', start_dt)
        event.add('dtend', end_dt)
        event.add('location', m.zoom_link or 'Online')
        event.add('description', f"{m.description or ''}\n\nChair: {m.chair_signup.display_name_snapshot + ' (' + m.chair_signup.user.bp_id + ')' if m.chair_signup else 'No chair yet'}")
        event.add('url', url_for('meeting_detail', meeting_id=m.id, _external=True))
        event.add('uid', f"meeting-{m.id}@backporchmeetings.org")
        
        cal.add_component(event)
    
    response = make_response(cal.to_ical())
    response.headers['Content-Type'] = 'text/calendar; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=backporch-calendar.ics'
    return response


@app.route("/my-calendar.ics")
@require_login
def my_calendar_ics():
    """Export user's committed meetings as iCal calendar feed."""
    user = get_current_user()
    if not user:
        abort(403)
    
    # Get meetings where this user is the chair
    my_meetings = (
        Meeting.query
        .join(ChairSignup)
        .filter(
            ChairSignup.user_id == user.id,
            Meeting.event_date >= date.today() - timedelta(days=30)  # Include recent past
        )
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .all()
    )
    
    cal = Calendar()
    cal.add('prodid', '-//Back Porch Meetings//backporchmeetings.org//')
    cal.add('version', '2.0')
    cal.add('x-wr-calname', f'My Back Porch Meetings - {user.display_name}')
    
    for m in my_meetings:
        event = Event()
        start_dt = datetime.combine(m.event_date, m.start_time)
        end_dt = datetime.combine(m.event_date, m.end_time) if m.end_time else start_dt + timedelta(hours=1)
        
        event.add('summary', f"Chair: {m.title}")
        event.add('dtstart', start_dt)
        event.add('dtend', end_dt)
        event.add('location', m.zoom_link or 'Online')
        event.add('description', f"You are chairing this meeting.\n\n{m.description or ''}")
        event.add('url', url_for('meeting_detail', meeting_id=m.id, _external=True))
        event.add('uid', f"my-meeting-{m.id}@backporchmeetings.org")
        
        # Add reminder 24 hours before
        alarm = Alarm()
        alarm.add('action', 'DISPLAY')
        alarm.add('description', f'Reminder: You are chairing {m.title} tomorrow')
        alarm.add('trigger', timedelta(hours=-24))
        event.add_component(alarm)
        
        # Add reminder 1 hour before
        alarm2 = Alarm()
        alarm2.add('action', 'DISPLAY')
        alarm2.add('description', f'Reminder: You are chairing {m.title} in 1 hour')
        alarm2.add('trigger', timedelta(hours=-1))
        event.add_component(alarm2)
        
        cal.add_component(event)
    
    response = make_response(cal.to_ical())
    response.headers['Content-Type'] = 'text/calendar; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=my-backporch-meetings.ics'
    return response


@app.route("/calendar/google-add/<int:meeting_id>")
@require_login
def google_calendar_add(meeting_id):
    """Generate Google Calendar add link for a specific meeting."""
    meeting = Meeting.query.get_or_404(meeting_id)
    user = get_current_user()
    
    # Only allow if user is the chair
    if not meeting.chair_signup or meeting.chair_signup.user_id != user.id:
        flash("You can only add meetings you're chairing to your calendar.", "warning")
        return redirect(url_for('meeting_detail', meeting_id=meeting_id))
    
    # Format for Google Calendar URL
    start_dt = datetime.combine(meeting.event_date, meeting.start_time)
    end_dt = datetime.combine(meeting.event_date, meeting.end_time) if meeting.end_time else start_dt + timedelta(hours=1)
    
    # Google Calendar date format: YYYYMMDDTHHMMSSZ
    start_str = start_dt.strftime('%Y%m%dT%H%M%S')
    end_str = end_dt.strftime('%Y%m%dT%H%M%S')
    
    title = f"Chair: {meeting.title}"
    details = f"You are chairing this Back Porch meeting.\n\n{meeting.description or ''}\n\nMeeting Link: {meeting.zoom_link or 'TBD'}"
    location = meeting.zoom_link or "Online"
    
    # Encode URL parameters
    from urllib.parse import quote
    
    google_url = (
        f"https://calendar.google.com/calendar/render?"
        f"action=TEMPLATE"
        f"&text={quote(title)}"
        f"&dates={start_str}/{end_str}"
        f"&details={quote(details)}"
        f"&location={quote(location)}"
        f"&sf=true&output=xml"
    )
    
    return redirect(google_url)


# ==========================
# CALENDAR INTEGRATION ROUTES
# ==========================

@app.route("/calendar/export")
@require_login
def calendar_export():
    """Calendar export options page."""
    user = get_current_user()
    
    # Get user's upcoming meetings
    upcoming_meetings = (
        Meeting.query
        .join(ChairSignup)
        .filter(
            ChairSignup.user_id == user.id,
            Meeting.event_date >= date.today()
        )
        .order_by(Meeting.event_date.asc())
        .limit(5)
        .all()
    )
    
    # Generate calendar subscription URLs
    ics_url = url_for('my_calendar_ics', _external=True)
    webcal_url = ics_url.replace('http://', 'webcal://').replace('https://', 'webcal://')
    
    return render_template(
        'calendar_export.html',
        upcoming_meetings=upcoming_meetings,
        ics_url=ics_url,
        webcal_url=webcal_url
    )
    response.headers['Content-Type'] = 'text/calendar; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=backporch-calendar.ics'
    return response


# ==========================
# AUTH ROUTES
# ==========================

@app.route("/register", methods=["GET", "POST"])
def register():
    # Registration feature gate
    if not app.config.get('REGISTRATION_ENABLED', True):
        flash("Registration is currently disabled. Please contact the admin.", "warning")
        return redirect(url_for("index"))
    
    # Check if user is logged in (graceful handling of DB issues)
    try:
        current_user = get_current_user()
        if current_user:
            flash("You are already logged in.", "info")
            return redirect(url_for("index"))
    except Exception:
        # If DB is unavailable, assume user is not logged in
        pass

    form = RegisterForm()
    if form.validate_on_submit():
        # Enforce access code(s) if configured
        provided_code = (form.access_code.data or '').strip()
        required_code = app.config.get('REGISTRATION_ACCESS_CODE')
        codes_list_raw = app.config.get('REGISTRATION_ACCESS_CODES')  # optional comma-separated list
        if required_code or codes_list_raw:
            valid_codes = set()
            if required_code:
                valid_codes.add(required_code.strip())
            if codes_list_raw:
                for c in str(codes_list_raw).split(','):
                    c = c.strip()
                    if c:
                        valid_codes.add(c)
            if provided_code not in valid_codes:
                flash("Invalid access code.", "danger")
                access_codes_configured = True
                return render_template("register.html", form=form, access_codes_configured=access_codes_configured)
        
        # Check for existing user
        try:
            existing = User.query.filter_by(email=form.email.data.lower().strip()).first()
        except Exception as e:
            app.logger.error(f"Database error checking existing user: {e}")
            flash("Database connection error. Please try again later or contact support.", "danger")
            access_codes_configured = bool(app.config.get('REGISTRATION_ACCESS_CODE') or app.config.get('REGISTRATION_ACCESS_CODES'))
            return render_template("register.html", form=form, access_codes_configured=access_codes_configured)
            
        if existing:
            flash("An account with that email already exists.", "danger")
            access_codes_configured = bool(app.config.get('REGISTRATION_ACCESS_CODE') or app.config.get('REGISTRATION_ACCESS_CODES'))
            return render_template("register.html", form=form, access_codes_configured=access_codes_configured)
        else:
            # Compute sobriety days from provided sobriety_date (optional)
            sob_days = None
            if form.sobriety_date.data:
                try:
                    d0 = form.sobriety_date.data
                    delta = (date.today() - d0).days
                    sob_days = max(0, delta)
                except Exception:
                    sob_days = None
            
            user = User(
                display_name=form.display_name.data.strip(),
                email=form.email.data.lower().strip(),
                is_admin=False,
                sobriety_days=sob_days,
                agreed_guidelines=form.agreed_guidelines.data,
                gender=form.gender.data,
            )
            user.set_password(form.password.data)
            
            try:
                db.session.add(user)
                db.session.commit()
                session["user_id"] = user.id
                app.logger.info(f"New user registered: {user.email} (ID: {user.id})")
                flash("Chairperson account created. Thank you for your service!", "success")
                next_url = request.args.get("next") or url_for("dashboard")
                return redirect(next_url)
            except Exception as e:
                db.session.rollback()
                # Check if this is a duplicate email constraint violation
                error_str = str(e).lower()
                if 'unique constraint failed' in error_str or 'duplicate entry' in error_str or 'duplicate key' in error_str:
                    app.logger.warning(f"Duplicate email registration attempt: {user.email}")
                    flash("An account with that email already exists.", "danger")
                else:
                    app.logger.error(f"Database error saving user registration: {e}")
                    flash("Registration failed due to database error. Please try again later or contact support.", "danger")
                access_codes_configured = bool(app.config.get('REGISTRATION_ACCESS_CODE') or app.config.get('REGISTRATION_ACCESS_CODES'))
                return render_template("register.html", form=form, access_codes_configured=access_codes_configured)
    # Provide flag to template so it can lock inputs until a code is entered
    access_codes_configured = bool(app.config.get('REGISTRATION_ACCESS_CODE') or app.config.get('REGISTRATION_ACCESS_CODES'))
    # Pass optional PDF names for responsibilities and protocol
    responsibilities_pdf = os.environ.get('CHAIR_RESPONSIBILITIES_PDF', 'Chairperson Responsibilities.pdf')
    protocol_pdf = os.environ.get('MEETING_PROTOCOL_PDF', 'Meeting Protocol.pdf')
    try:
        # Build dynamic list of available PDFs from configured folder
        pdfs = []
        try:
            if os.path.isdir(EXTERNAL_PDFS_DIR):
                for name in sorted(os.listdir(EXTERNAL_PDFS_DIR)):
                    if name.lower().endswith('.pdf') and not name.startswith('.'):
                        pdfs.append({
                            'name': name,
                            'url': url_for('serve_pdf', filename=name)
                        })
        except Exception:
            pdfs = []

        return render_template(
            "register.html",
            form=form,
            access_codes_configured=access_codes_configured,
            responsibilities_pdf=responsibilities_pdf,
            protocol_pdf=protocol_pdf,
            pdfs=pdfs,
        )
    except Exception as e:
        import traceback
        print("Register page render error:")
        print(traceback.format_exc())
        return f"Register page error: {e}", 500


@app.route("/login", methods=["GET", "POST"])
def login():
    if get_current_user():
        flash("You are already logged in.", "info")
        return redirect(url_for("index"))

    def _strip_zero_width(s: str) -> str:
        # Common invisible characters introduced by mobile keyboards / copy-paste.
        return (
            (s or "")
            .replace("\u200b", "")  # zero width space
            .replace("\u200c", "")  # zero width non-joiner
            .replace("\u200d", "")  # zero width joiner
            .replace("\ufeff", "")  # BOM
        )

    def _normalize_email(raw: str) -> str:
        return _strip_zero_width((raw or "").strip().lower())

    form = LoginForm()
    if form.validate_on_submit():
        email = _normalize_email(form.email.data)
        user = User.query.filter_by(email=email).first()

        # Security configuration (configurable to reduce mobile lockout pain)
        enable_lockout = bool(app.config.get('ENABLE_ACCOUNT_LOCKOUT', True))
        max_attempts = int(app.config.get('MAX_FAILED_LOGIN_ATTEMPTS', 10) or 0)
        lock_minutes = int(app.config.get('LOCKOUT_MINUTES', 10) or 0)
        allow_trim = bool(app.config.get('ALLOW_PASSWORD_TRIM_ON_LOGIN', True))
        allow_locked_login = bool(app.config.get('ALLOW_LOCKED_ACCOUNT_LOGIN_IF_PASSWORD_CORRECT', True))
        
        if user:
            raw_password = form.password.data or ""
            # Build candidate passwords to tolerate common mobile autofill quirks.
            # We never store or log the password itself.
            candidates = [raw_password]
            normalized_pw = _strip_zero_width(raw_password)
            if normalized_pw != raw_password:
                candidates.append(normalized_pw)
            if allow_trim:
                stripped_pw = raw_password.strip()
                if stripped_pw and stripped_pw != raw_password:
                    candidates.append(stripped_pw)
                stripped_normalized_pw = normalized_pw.strip()
                if stripped_normalized_pw and stripped_normalized_pw not in candidates:
                    candidates.append(stripped_normalized_pw)

            password_ok_direct = user.check_password(raw_password)
            password_ok = password_ok_direct or any(user.check_password(p) for p in candidates[1:])

            # If account is locked, allow login if the correct password is provided.
            if user.is_locked() and not (allow_locked_login and password_ok):
                log_audit_event('login_attempt_locked_account', user.id, details={
                    'email': email,
                    'failed_attempts': user.failed_login_attempts,
                    'locked_until': user.locked_until.isoformat() if user.locked_until else None,
                    'user_agent': (request.headers.get('User-Agent') or '')[:200],
                    'password_length': len(raw_password),
                    'password_had_outer_whitespace': raw_password != raw_password.strip(),
                })
                flash("Account is temporarily locked due to too many failed login attempts. Please try again later.", "danger")
                return render_template("login.html", form=form)

            # Check password
            if password_ok:
                # Successful login
                session.permanent = True  # Use PERMANENT_SESSION_LIFETIME from config
                session["user_id"] = user.id
                session.modified = True  # Force session to be saved
                user.last_login = datetime.utcnow()
                user.failed_login_attempts = 0  # Reset failed attempts
                user.locked_until = None
                db.session.commit()
                
                log_audit_event('login_success', user.id, details={
                    'email': email,
                    'user_agent': (request.headers.get('User-Agent') or '')[:200],
                    'compat_password_normalization_used': bool(allow_trim and (raw_password != raw_password.strip() or raw_password != _strip_zero_width(raw_password))),
                    'session_id_set': bool(session.get('user_id')),
                })

                if allow_trim and (not password_ok_direct) and raw_password != raw_password.strip():
                    # Gentle hint: this is a common mobile autofill issue.
                    flash(
                        "Note: your password appeared to include extra spaces. If you use autofill on mobile, double-check it doesn’t add a trailing space.",
                        "warning",
                    )
                
                # Check if password reset is required
                if user.password_reset_required:
                    flash("You must change your password before continuing.", "warning")
                    return redirect(url_for("change_password", required=True))
                
                flash("Logged in successfully.", "success")
                next_url = request.args.get("next") or url_for("index")
                return redirect(next_url)
            else:
                # Failed login - increment attempts
                user.failed_login_attempts += 1

                should_lock = False
                if enable_lockout and max_attempts > 0 and user.failed_login_attempts >= max_attempts:
                    should_lock = True

                if should_lock and lock_minutes > 0:
                    user.lock_account(lock_minutes)
                    log_audit_event('account_locked', user.id, details={
                        'email': email,
                        'failed_attempts': user.failed_login_attempts,
                        'lock_minutes': lock_minutes,
                        'user_agent': (request.headers.get('User-Agent') or '')[:200],
                        'password_length': len(raw_password),
                        'password_had_outer_whitespace': raw_password != raw_password.strip(),
                    })
                    flash(f"Account locked due to too many failed login attempts. Please try again in {lock_minutes} minutes.", "danger")
                else:
                    db.session.commit()
                    log_audit_event('login_failure', user.id, details={
                        'email': email,
                        'failed_attempts': user.failed_login_attempts,
                        'max_attempts': max_attempts if enable_lockout else None,
                        'user_agent': (request.headers.get('User-Agent') or '')[:200],
                        'password_length': len(raw_password),
                        'password_had_outer_whitespace': raw_password != raw_password.strip(),
                    })
                    if enable_lockout and max_attempts > 0:
                        remaining = max(0, max_attempts - user.failed_login_attempts)
                        flash(f"Invalid password. {remaining} attempts remaining.", "danger")
                    else:
                        flash("Invalid email or password.", "danger")
        else:
            # User not found
            log_audit_event('login_attempt_invalid_user', details={
                'email': email,
                'user_agent': (request.headers.get('User-Agent') or '')[:200],
            })
            flash("Invalid email or password.", "danger")
            
    return render_template("login.html", form=form)


@app.route("/logout")
def logout():
    user = get_current_user()
    if user:
        log_audit_event('logout', user.id)
    
    session.pop("user_id", None)
    flash("Logged out.", "info")
    return redirect(url_for("index"))

# API: validate registration unlock key
@app.route("/api/registration/validate-key", methods=["POST"])
def api_validate_registration_key():
    try:
        data = request.get_json(force=True) or {}
        provided = (data.get('key') or '').strip()
        required_code = app.config.get('REGISTRATION_ACCESS_CODE')
        codes_list_raw = app.config.get('REGISTRATION_ACCESS_CODES')
        valid_codes = set()
        if required_code:
            valid_codes.add(str(required_code).strip())
        if codes_list_raw:
            for c in str(codes_list_raw).split(','):
                c = c.strip()
                if c:
                    valid_codes.add(c)
        ok = True
        if valid_codes:
            ok = provided in valid_codes
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ==========================
# USER PAGES (secured)
# ==========================

@app.route("/dashboard", methods=["GET", "POST"])
@login_required
def dashboard():
    try:
        user = get_current_user()
        today = date.today()
        
        # Handle profile form submission
        form = ProfileForm(obj=user)
        if form.validate_on_submit():
            user.display_name = form.display_name.data.strip()
            user.gender = form.gender.data or None
            db.session.commit()
            flash("Profile updated successfully!", "success")
            return redirect(url_for("dashboard"))
        
        # Get user's meetings - use explicit SELECT with proper joins
        # Only get meetings where ChairSignup.user_id matches current user
        all_meetings = (
            db.session.query(Meeting)
            .select_from(Meeting)
            .join(ChairSignup, Meeting.id == ChairSignup.meeting_id)
            .filter(ChairSignup.user_id == user.id)
            .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
            .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
            .all()
        )
        
        # Debug logging
        app.logger.info(f"Dashboard for user {user.id} ({user.email}): Found {len(all_meetings)} total meetings")
        for m in all_meetings:
            app.logger.info(f"  - Meeting {m.id}: {m.title} on {m.event_date}")
        
        # Separate today's, upcoming, and past meetings
        todays_meetings = [m for m in all_meetings if m.event_date == today]
        upcoming_meetings = [m for m in all_meetings if m.event_date > today]
        past_meetings = [m for m in all_meetings if m.event_date < today]
        
        # Combine today's and future meetings for "upcoming commitments" stat
        all_future_meetings = todays_meetings + upcoming_meetings
        
        # Get next upcoming meeting (could be today or future)
        next_meeting = all_future_meetings[0] if all_future_meetings else None
        
        # Get user's availability signups
        try:
            upcoming_availability = (
                ChairpersonAvailability.query
                .filter_by(user_id=user.id, is_active=True)
                .filter(ChairpersonAvailability.volunteer_date >= today)
                .order_by(ChairpersonAvailability.volunteer_date.asc())
                .limit(5)
                .all()
            )
        except Exception as e:
            app.logger.error(f"Error getting availability: {e}")
            upcoming_availability = []
        
        # Calculate service stats
        total_meetings_chaired = len(past_meetings)
        upcoming_commitments = len(all_future_meetings)  # Include today's meetings
        volunteer_signups = len(upcoming_availability)
        
        # Calculate recent meetings (last 30 days)
        recent_cutoff = today - timedelta(days=30)
        recent_meetings = [m for m in past_meetings if m.event_date >= recent_cutoff]
        recent_meetings_count = len(recent_meetings)
        
        # Get recent open meetings that need chairs
        try:
            open_meetings = (
                Meeting.query
                .filter(
                    Meeting.event_date >= today,
                    Meeting.event_date <= today + timedelta(days=30),
                    Meeting.is_open == True,
                    ~Meeting.chair_signup.has()
                )
                .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
                .limit(5)
                .all()
            )
        except Exception as e:
            app.logger.error(f"Error getting open meetings: {e}")
            open_meetings = []
        
        # Calculate days until next meeting
        days_until_next = None
        if next_meeting:
            days_until_next = (next_meeting.event_date - today).days
        
        # Admin statistics (only calculate if user is admin)
        admin_stats = None
        admin_charts = None
        if user.is_admin:
            try:
                # Count all registered hosts (non-admin users)
                total_hosts = User.query.filter_by(is_admin=False, is_active=True).count()
                
                # Count current chairs (meetings from today going forward that have chair signups)
                current_chairs = (
                    db.session.query(Meeting)
                    .filter(
                        Meeting.event_date >= today,
                        Meeting.chair_signup.has()
                    )
                    .count()
                )
                
                # Count future chairs needed (meetings from today going forward)
                future_chairs_needed = (
                    Meeting.query
                    .filter(Meeting.event_date >= today)
                    .count()
                )
                
                # Calculate percentage of completed chairs
                chair_completion_percent = 0
                if future_chairs_needed > 0:
                    chair_completion_percent = round((current_chairs / future_chairs_needed) * 100, 1)
                
                admin_stats = {
                    'total_hosts': total_hosts,
                    'current_chairs': current_chairs,
                    'future_chairs_needed': future_chairs_needed,
                    'chair_completion_percent': chair_completion_percent,
                    'unfilled_chairs': future_chairs_needed - current_chairs
                }
                
                # Calculate chart data for last 30 days
                last_30_days = today - timedelta(days=30)
                
                # Men's meetings coverage (last 30 days)
                mens_total = Meeting.query.filter(
                    Meeting.event_date >= last_30_days,
                    Meeting.event_date <= today,
                    Meeting.gender_restriction == 'men'
                ).count()
                
                mens_filled = db.session.query(Meeting).filter(
                    Meeting.event_date >= last_30_days,
                    Meeting.event_date <= today,
                    Meeting.gender_restriction == 'men',
                    Meeting.chair_signup.has()
                ).count()
                
                # Women's meetings coverage (last 30 days)
                womens_total = Meeting.query.filter(
                    Meeting.event_date >= last_30_days,
                    Meeting.event_date <= today,
                    Meeting.gender_restriction == 'women'
                ).count()
                
                womens_filled = db.session.query(Meeting).filter(
                    Meeting.event_date >= last_30_days,
                    Meeting.event_date <= today,
                    Meeting.gender_restriction == 'women',
                    Meeting.chair_signup.has()
                ).count()
                
                # All meetings coverage (last 30 days)
                all_total = Meeting.query.filter(
                    Meeting.event_date >= last_30_days,
                    Meeting.event_date <= today
                ).count()
                
                all_filled = db.session.query(Meeting).filter(
                    Meeting.event_date >= last_30_days,
                    Meeting.event_date <= today,
                    Meeting.chair_signup.has()
                ).count()
                
                # Weekly trend data (last 4 weeks)
                weekly_data = []
                for week in range(4):
                    week_start = today - timedelta(days=(week + 1) * 7)
                    week_end = today - timedelta(days=week * 7)
                    
                    week_total = Meeting.query.filter(
                        Meeting.event_date >= week_start,
                        Meeting.event_date < week_end
                    ).count()
                    
                    week_filled = db.session.query(Meeting).filter(
                        Meeting.event_date >= week_start,
                        Meeting.event_date < week_end,
                        Meeting.chair_signup.has()
                    ).count()
                    
                    weekly_data.insert(0, {
                        'week': f'Week {4-week}',
                        'total': week_total,
                        'filled': week_filled,
                        'percent': round((week_filled / week_total * 100) if week_total > 0 else 0, 1)
                    })
                
                # User participation percentages (top 10)
                user_participation = []
                users = User.query.filter_by(is_admin=False).all()
                
                for u in users:
                    user_total = db.session.query(Meeting).join(ChairSignup).filter(
                        ChairSignup.user_id == u.id
                    ).count()
                    
                    if user_total > 0:
                        user_participation.append({
                            'name': u.display_name,
                            'count': user_total,
                            'percent': round((user_total / all_total * 100) if all_total > 0 else 0, 1)
                        })
                
                user_participation.sort(key=lambda x: x['count'], reverse=True)
                user_participation = user_participation[:10]
                
                admin_charts = {
                    'mens_coverage': {
                        'filled': mens_filled,
                        'unfilled': mens_total - mens_filled,
                        'percent': round((mens_filled / mens_total * 100) if mens_total > 0 else 0, 1)
                    },
                    'womens_coverage': {
                        'filled': womens_filled,
                        'unfilled': womens_total - womens_filled,
                        'percent': round((womens_filled / womens_total * 100) if womens_total > 0 else 0, 1)
                    },
                    'overall_coverage': {
                        'filled': all_filled,
                        'unfilled': all_total - all_filled,
                        'percent': round((all_filled / all_total * 100) if all_total > 0 else 0, 1)
                    },
                    'weekly_trend': weekly_data,
                    'user_participation': user_participation
                }
                
                app.logger.info(f"Admin stats: {admin_stats}")
                app.logger.info(f"Admin charts: {admin_charts}")
                
            except Exception as e:
                app.logger.error(f"Error calculating admin stats: {e}")
                import traceback
                traceback.print_exc()
                admin_stats = None
                admin_charts = None
        
        return render_template(
            "dashboard.html",
            user=user,
            form=form,
            today=today,
            todays_meetings=todays_meetings,
            upcoming_meetings=upcoming_meetings[:5],
            past_meetings=past_meetings[-10:],
            next_meeting=next_meeting,
            days_until_next=days_until_next,
            upcoming_availability=upcoming_availability,
            total_meetings_chaired=total_meetings_chaired,
            upcoming_commitments=upcoming_commitments,
            volunteer_signups=volunteer_signups,
            recent_meetings_count=recent_meetings_count,
            open_meetings=open_meetings,
            admin_stats=admin_stats,
            admin_charts=admin_charts
        )
    except Exception as e:
        app.logger.error(f"Dashboard error for user {user.id if user else 'unknown'}: {e}")
        import traceback
        traceback.print_exc()
        flash("Error loading dashboard. Please try again or contact support.", "danger")
        return redirect(url_for("index"))


@app.route("/dashboard/refresh")
@login_required
def dashboard_refresh():
    """Refresh dashboard by clearing cache and reloading data."""
    try:
        # Clear all cached data
        try:
            cache.clear()
        except Exception as cache_error:
            app.logger.warning(f"Failed to clear cache: {cache_error}")
        flash("Dashboard data refreshed!", "success")
    except Exception as e:
        app.logger.error(f"Error refreshing dashboard: {e}")
        flash("Error refreshing data. Please try again.", "warning")
    
    return redirect(url_for("dashboard"))


@app.route("/debug/my-signups")
@login_required
def debug_my_signups():
    """Debug route to check user's chair signups."""
    user = get_current_user()
    
    # Get all ChairSignup records for this user
    signups = ChairSignup.query.filter_by(user_id=user.id).all()
    
    # Get all meetings through the join - using same query as dashboard
    meetings_query = (
        db.session.query(Meeting)
        .select_from(Meeting)
        .join(ChairSignup, Meeting.id == ChairSignup.meeting_id)
        .filter(ChairSignup.user_id == user.id)
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
    )
    
    # Get the compiled SQL for debugging
    from sqlalchemy.dialects import mysql
    sql_text = str(meetings_query.statement.compile(dialect=mysql.dialect(), compile_kwargs={"literal_binds": True}))
    
    meetings = meetings_query.all()
    
    debug_info = {
        "user_id": user.id,
        "user_email": user.email,
        "user_display_name": user.display_name,
        "total_signups": len(signups),
        "total_meetings": len(meetings),
        "sql_query": sql_text,
        "signups": [
            {
                "id": s.id,
                "meeting_id": s.meeting_id,
                "display_name": s.display_name_snapshot,
                "created_at": s.created_at.isoformat() if s.created_at else None,
                "meeting_exists": Meeting.query.get(s.meeting_id) is not None
            }
            for s in signups
        ],
        "meetings": [
            {
                "id": m.id,
                "title": m.title,
                "event_date": m.event_date.isoformat(),
                "start_time": m.start_time.strftime('%H:%M:%S') if m.start_time else None,
                "has_chair": m.has_chair,
                "chair_signup_id": m.chair_signup.id if m.chair_signup else None,
                "chair_user_id": m.chair_signup.user_id if m.chair_signup else None
            }
            for m in meetings
        ]
    }
    
    return jsonify(debug_info)


@cache.memoize(timeout=180)  # Cache for 3 minutes
def get_open_meetings_cached():
    """Get open meetings that need chairs - cached for performance."""
    today = date.today()
    
    return (
        Meeting.query
        .filter(
            Meeting.event_date >= today,
            Meeting.event_date <= today + timedelta(days=30),
            Meeting.is_open == True,
            ~Meeting.chair_signup.has()
        )
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .limit(20)  # Reasonable limit
        .all()
    )


@cache.memoize(timeout=600)  # Cache for 10 minutes
def get_meeting_stats_cached():
    """Get meeting statistics - cached for performance."""
    today = date.today()
    
    # Use more efficient count queries
    total_meetings = Meeting.query.count()
    upcoming_meetings = Meeting.query.filter(Meeting.event_date >= today).count()
    need_chairs = Meeting.query.filter(
        Meeting.event_date >= today,
        Meeting.is_open == True,
        ~Meeting.chair_signup.has()
    ).count()
    
    return {
        'total_meetings': total_meetings,
        'upcoming_meetings': upcoming_meetings,
        'need_chairs': need_chairs
    }


@cache.memoize(timeout=1800)  # Cache for 30 minutes
def get_analytics_data_cached():
    """Get analytics data for admin dashboard - cached for performance."""
    # This would contain the heavy analytics queries
    # Moved from admin_analytics route for better performance
    pass


# Cache busting function for when data changes
def clear_dashboard_cache(user_id=None):
    """Clear dashboard cache when meetings change."""
    if user_id:
        # Clear specific user's cache
        today = date.today()
        cache.delete(f"dashboard_data_{user_id}_{today.isoformat()}")
    else:
        # Clear all dashboard caches (less efficient but thorough)
        cache.clear()


# Cache busting when meetings are modified
def invalidate_meeting_caches():
    """Invalidate all meeting-related caches."""
    cache.delete_memoized(get_open_meetings_cached)
    cache.delete_memoized(get_meeting_stats_cached)
    cache.delete_memoized(get_analytics_data_cached)


class ProfileForm(FlaskForm):
    display_name = StringField("Display Name", validators=[DataRequired(), Length(max=80)])
    gender = RadioField(
        "Gender",
        choices=[('male','Male'),('female','Female')],
        validators=[Optional()]
    )
    profile_image = FileField("Profile Image", validators=[Optional()])
    submit = SubmitField("Save Changes")


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    try:
        user = get_current_user()
        if not user:
            flash("Please log in to view your profile.", "warning")
            return redirect(url_for("login"))
        
        form = ProfileForm(obj=user)
        
        # Get search and filter parameters for meeting history
        search_query = request.args.get('search', '').strip()
        meeting_type_filter = request.args.get('meeting_type', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        show_section = request.args.get('show', 'upcoming')  # upcoming, past, availability
    
        # Base query for user's chair signups
        base_query = (
            ChairSignup.query
            .filter_by(user_id=user.id)
            .join(Meeting)
            .options(db.joinedload(ChairSignup.meeting))
            )
    
        # Apply search filter to meetings
        if search_query:
            base_query = base_query.filter(
                db.or_(
                    Meeting.title.ilike(f'%{search_query}%'),
                    Meeting.description.ilike(f'%{search_query}%')
                )
            )
        
        # Apply meeting type filter
        if meeting_type_filter:
            base_query = base_query.filter(Meeting.meeting_type == meeting_type_filter)
        
        # Apply date range filters
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                base_query = base_query.filter(Meeting.event_date >= from_date)
            except ValueError:
                pass
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                base_query = base_query.filter(Meeting.event_date <= to_date)
            except ValueError:
                pass
        
        # Get all filtered chair signups
        chair_signups = base_query.order_by(Meeting.event_date.asc(), Meeting.start_time.asc()).all()
        
        # Get user's availability signups (volunteer dates)
        availability_query = ChairpersonAvailability.query.filter_by(user_id=user.id, is_active=True)
        
        # Apply date range filters to availability if provided
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                availability_query = availability_query.filter(ChairpersonAvailability.volunteer_date >= from_date)
            except ValueError:
                pass
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                availability_query = availability_query.filter(ChairpersonAvailability.volunteer_date <= to_date)
            except ValueError:
                pass
        
        availability_signups = availability_query.order_by(ChairpersonAvailability.volunteer_date.asc()).all()
        
        # Separate past and future meetings
        from datetime import date
        today = date.today()
        
        past_meetings = [signup for signup in chair_signups if signup.meeting.event_date < today]
        upcoming_meetings = [signup for signup in chair_signups if signup.meeting.event_date >= today]
        
        # Separate past and future availability
        past_availability = [avail for avail in availability_signups if avail.volunteer_date < today]
        upcoming_availability = [avail for avail in availability_signups if avail.volunteer_date >= today]
        
        # Get unique meeting types for filter dropdown
        user_meeting_types = db.session.query(Meeting.meeting_type).join(ChairSignup).filter(ChairSignup.user_id == user.id).distinct().all()
        user_meeting_types = [mt[0] for mt in user_meeting_types if mt[0]]
        
        if form.validate_on_submit():
            try:
                user.display_name = form.display_name.data.strip()
                user.gender = form.gender.data or None
                
                # Handle profile image upload - store as base64 in database
                if form.profile_image.data:
                    file = form.profile_image.data
                    if file and allowed_file(file.filename):
                        # Read image data
                        image_data = file.read()
                        # Validate image size (max 5MB)
                        if len(image_data) > 5 * 1024 * 1024:
                            flash("Image file is too large. Maximum size is 5MB.", "danger")
                            return redirect(url_for("profile"))
                        # Encode to base64 and store as text
                        user.profile_image = base64.b64encode(image_data).decode('utf-8')
                
                db.session.commit()
                flash("Profile updated.", "success")
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"Error updating profile: {e}")
                import traceback
                app.logger.error(traceback.format_exc())
                flash(f"Error updating profile: {str(e)}", "danger")
            return redirect(url_for("profile"))
    
        # Build current filters for template
        current_filters = {
            'search': search_query,
            'meeting_type': meeting_type_filter,
            'date_from': date_from,
            'date_to': date_to
        }
        
        return render_template("profile.html", 
                             form=form, 
                             user=user,
                             upcoming_meetings=upcoming_meetings,
                             past_meetings=past_meetings,
                             upcoming_availability=upcoming_availability,
                             past_availability=past_availability,
                             user_meeting_types=user_meeting_types,
                             current_filters=current_filters)
    except Exception as e:
        app.logger.error(f"Error loading profile page: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        flash(f"Error loading profile: {str(e)}", "danger")
        return redirect(url_for("dashboard"))


@app.route("/profile/refresh")
@login_required
def profile_refresh():
    """Refresh profile by clearing cache and reloading data."""
    try:
        # Clear all cached data
        try:
            cache.clear()
        except Exception as cache_error:
            app.logger.warning(f"Failed to clear cache: {cache_error}")
        flash("Profile data refreshed!", "success")
    except Exception as e:
        app.logger.error(f"Error refreshing profile: {e}")
        flash("Error refreshing data. Please try again.", "warning")
    
    return redirect(url_for("profile"))


@app.route("/profile/image/<int:user_id>")
def profile_image(user_id):
    """Serve profile image from database."""
    user = User.query.get_or_404(user_id)
    if user.profile_image:
        try:
            # Decode base64 image data
            image_data = base64.b64decode(user.profile_image)
            # Default to jpeg since we're not storing type
            return Response(image_data, mimetype='image/jpeg')
        except Exception as e:
            app.logger.error(f"Error decoding profile image for user {user_id}: {e}")
            return "", 404
    else:
        # Return a default placeholder image or 404
        return "", 404


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    """Allow user to change their password."""
    user = get_current_user()
    if not user:
        flash("Please log in.", "warning")
        return redirect(url_for("login"))
    
    required = request.args.get('required') == 'True'
    
    if request.method == "POST":
        current_password = request.form.get("current_password")
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")
        
        # Validation
        if not all([current_password, new_password, confirm_password]):
            flash("All fields are required.", "danger")
            return render_template("change_password.html", required=required)
        
        # Check current password (skip if password reset is required)
        if not user.password_reset_required and not user.check_password(current_password):
            flash("Current password is incorrect.", "danger")
            return render_template("change_password.html", required=required)
        
        # Check new passwords match
        if new_password != confirm_password:
            flash("New passwords do not match.", "danger")
            return render_template("change_password.html", required=required)
        
        # Check password strength
        if len(new_password) < 8:
            flash("Password must be at least 8 characters long.", "danger")
            return render_template("change_password.html", required=required)
        
        # Update password
        user.set_password(new_password)
        user.password_reset_required = False  # Clear the reset flag
        db.session.commit()
        
        log_audit_event('password_changed', user.id)
        flash("Password changed successfully!", "success")
        return redirect(url_for("dashboard"))
    
    return render_template("change_password.html", required=required)


# ==========================
# VIDEO QUIZ SYSTEM
# ==========================

# Quiz data structure
QUIZZES = {
    'registration': {
        'id': 'registration',
        'title': 'How to Register to Host a BP Online Zoom Meeting',
        'video': 'BP_Zoom_Host_Registration_Walkthrough.mp4',
        'description': 'Learn how to register as a chairperson for Back Porch meetings.',
        'questions': [
            {
                'question': 'Who presents the tutorial in the transcript?',
                'options': ['John the guide', 'Jeff the genius', 'The Back Porch Committee', 'Zoom Tech Support'],
                'correct': 1
            },
            {
                'question': 'What is the main purpose of this tutorial?',
                'options': ['To show how to host in-person AA meetings', 'To explain how to access and register for the Back Porch Zoom host application', 'To teach advanced Zoom security settings', 'To review AA literature readings'],
                'correct': 1
            },
            {
                'question': 'According to the tutorial, how can you access the Back Porch Zoom host registration application?',
                'options': ['Only from the Zoom mobile app', 'Through a direct URL or by visiting therealbackporch.com', 'Only through an email invite', 'Through the AA World Services website'],
                'correct': 1
            },
            {
                'question': 'After reaching the Back Porch website, which link should you select to get to the sign-up portal?',
                'options': ['Meetings', 'Resources', 'Chairs', 'Contact'],
                'correct': 2
            },
            {
                'question': 'What does the green button on the Chairs page do?',
                'options': ['Logs you into your existing account', 'Opens the portal where you register as a host for Back Porch Zoom AA meetings', 'Starts a Zoom meeting immediately', 'Sends an email to support'],
                'correct': 1
            },
            {
                'question': 'What color is the Chairperson Sign Up button at the top of the application?',
                'options': ['Blue', 'Red', 'Yellow', 'Purple'],
                'correct': 2
            },
            {
                'question': 'What do you need in order to become a host and unlock the registration form?',
                'options': ['A Zoom Pro license', "A sponsor's approval code", 'A registration key', 'A host meeting ID'],
                'correct': 2
            },
            {
                'question': 'What happens when you select the link under the registration unlock key field?',
                'options': ['It opens a live chat with support', 'It generates an email that you send to Back Porch support to request your key', 'It automatically approves you as a host', 'It shows a list of available meetings'],
                'correct': 1
            },
            {
                'question': 'After entering your registration key and unlocking the form, which pieces of information are required?',
                'options': ['Name, phone number, home group, sponsor', 'Name, email, password, sobriety date', 'Name, address, Zoom ID, phone number', 'Email, meeting ID, password, phone'],
                'correct': 1
            },
            {
                'question': 'Why is your sobriety date required on the registration form?',
                'options': ['To send birthday announcements', 'To confirm that you meet the 90-day requirement to host', 'To create your Back Porch username', 'To share with other members on the calendar'],
                'correct': 1
            },
            {
                'question': 'How is gender selection used in the registration process?',
                'options': ['It is required for all meetings', 'It is used for marketing surveys', 'It is optional and used only for gender-specific meetings', 'It is used to decide your hosting level'],
                'correct': 2
            },
            {
                'question': 'Once your account is created, what are you instructed to do next?',
                'options': ['Wait for a phone call from support', 'Immediately choose a meeting date', 'Return to the home page, select Login, and enter your new credentials', 'Re-enter your sobriety date for verification'],
                'correct': 2
            },
            {
                'question': 'Which of the following is NOT mentioned as something you can do or see on your dashboard?',
                'options': ['Update your profile and upload a photo', 'View upcoming meetings you are scheduled to chair', 'Check your last 30 days of hosting history', 'Change the 90-day sobriety requirement'],
                'correct': 3
            },
            {
                'question': 'What must be true before you can register for a meeting on the calendar?',
                'options': ['You must be on a mobile device', 'You must be signed up and logged in', 'You must have hosted at least one meeting already', 'You must be approved by two other hosts'],
                'correct': 1
            },
            {
                'question': 'How are available dates displayed on the calendar?',
                'options': ['In blue with the label "Open"', 'In green with the label "Available"', 'In gray with the label "No Chair"', 'In yellow with the label "Unassigned"'],
                'correct': 2
            },
            {
                'question': 'After you select a date to chair a meeting, which emails does the system send?',
                'options': ['A confirmation email and a weekly newsletter', 'A confirmation email and a reminder email 24 hours before your scheduled meeting', 'Only a reminder email one hour before the meeting', 'Only a confirmation email if it\'s your first time hosting'],
                'correct': 1
            },
            {
                'question': 'If something changes and you cannot chair a meeting, what should you do?',
                'options': ['Ignore the meeting and let it pass', 'Email every group member', 'Go to your profile, select the red button next to the meeting to cancel, and confirm your choice', 'Call Zoom support to cancel'],
                'correct': 2
            },
            {
                'question': 'Who can view the Back Porch hosting calendar?',
                'options': ['Only registered hosts', 'Only members with 90 days or more of sobriety', 'Anyone can view it without registering', 'Only the web administrator'],
                'correct': 2
            },
            {
                'question': 'Which of the following is specifically mentioned as part of the "How to Chair Resources" section?',
                'options': ['Host payment policies', 'Zoom best practices and handling online disruptions', 'Step study handouts from other groups', 'Insurance and liability information'],
                'correct': 1
            },
            {
                'question': 'Which interface or display options are mentioned for viewing and using the calendar and profile?',
                'options': ['Sorting by region and switching to printer mode', 'Filtering by date or your name, and choosing between light mode and dark mode using the lamp icon', 'Enabling slideshow mode and auto-play', 'Switching between desktop and TV mode'],
                'correct': 1
            }
        ]
    },
    'hosting': {
        'id': 'hosting',
        'title': 'How to Host a Zoom Meeting',
        'video': 'BP_Zoom_Hosting_Chairing_Guide.mp4',
        'description': 'Learn the best practices for hosting a Back Porch Zoom meeting.',
        'questions': [
            {
                'question': 'What is the first step to start hosting a Back Porch Zoom meeting?',
                'options': ['Open Zoom and wait for participants', 'Visit therealbackporch.com/meetings and click the Zoom link', 'Enter the host code in Zoom settings', 'Start a breakout room'],
                'correct': 1
            },
            {
                'question': 'How do you claim host controls once the Zoom window opens?',
                'options': ['Click Security > Claim Host', 'Click Participants > Claim Host', 'Click Chat > Claim Host', 'Use the breakout room panel'],
                'correct': 1
            },
            {
                'question': 'What must you enter in order to become the host of the meeting?',
                'options': ['Your sobriety date', 'Your Back Porch login password', 'The host code emailed to you', 'Your personal meeting ID'],
                'correct': 2
            },
            {
                'question': 'What appears beside your name once you have successfully claimed host control?',
                'options': ['Admin', 'Owner', 'Host', 'Leader'],
                'correct': 2
            },
            {
                'question': 'What is the purpose of enabling the Waiting Room?',
                'options': ['To create breakout rooms', 'To prevent uninvited attendees from entering directly', 'To mute participants automatically', 'To assign chair points'],
                'correct': 1
            },
            {
                'question': 'How do you enable the Waiting Room?',
                'options': ['From the Chat panel', 'From the Security button', 'From the three dots at the bottom of the participant list', 'From the Audio settings'],
                'correct': 2
            },
            {
                'question': 'How do you make someone a co-host?',
                'options': ['Click the participant\'s name and choose "Make co-host"', 'Send them the host key', 'Promote them in the chat', 'Add them in Zoom settings before the meeting'],
                'correct': 0
            },
            {
                'question': 'Why should hosts mute all participants before the meeting begins?',
                'options': ['To reduce bandwidth', 'To ensure no one speaks out of turn', 'To activate screen sharing', 'To start recording automatically'],
                'correct': 1
            },
            {
                'question': 'Which setting must be unchecked to prevent participants from unmuting themselves?',
                'options': ['Allow participants to chat', 'Allow participants to rename', 'Allow participants to unmute themselves', 'Allow participants to start video'],
                'correct': 2
            },
            {
                'question': 'What should the host do if the chat contains offensive behavior?',
                'options': ['Report the person to Zoom headquarters', 'Delete the message using the three-dot menu', 'Transfer the host role', 'Remove all participants'],
                'correct': 1
            },
            {
                'question': 'How can the host turn off chat entirely?',
                'options': ['Click Security and uncheck Chat', 'Click Participants and uncheck Chat', 'Click Video and uncheck Chat', 'Click View and disable Chat'],
                'correct': 0
            },
            {
                'question': 'What happens once chat is turned off?',
                'options': ['It becomes read-only', 'It becomes visible only to hosts', 'The chat box disappears for the rest of the meeting', 'It restarts after 60 seconds'],
                'correct': 2
            },
            {
                'question': 'Where do you remove a disruptive participant?',
                'options': ['Breakout Rooms > Remove', 'Security > Kick Out', 'Manage Participants > More > Remove', 'Chat > Block User'],
                'correct': 2
            },
            {
                'question': 'What happens when you remove a disruptive attendee?',
                'options': ['They can re-enter at any time', 'They cannot return under the same account', 'They are permanently banned from Zoom', 'They are placed on hold in the waiting room'],
                'correct': 1
            },
            {
                'question': 'How do you stop a participant\'s video?',
                'options': ['In Security, uncheck Start Video', 'In Manage Participants, click More > Stop Video', 'In Chat, click Disable Camera', 'In Settings, turn off participant video'],
                'correct': 1
            },
            {
                'question': 'Which setting prevents participants from renaming themselves?',
                'options': ['Disable video', 'Uncheck Allow participants to rename themselves', 'Lock the meeting', 'Turn off screen sharing'],
                'correct': 1
            },
            {
                'question': 'How can the host restrict chat so participants can send messages only to the host?',
                'options': ['Use Chat > Private Mode', 'Set chat permissions to "Host Only"', 'Turn off all chat features', 'Add participants to a quiet list'],
                'correct': 1
            },
            {
                'question': 'What is the purpose of locking the meeting?',
                'options': ['To hide the participant list', 'To prevent new attendees from entering', 'To disable screen sharing', 'To reset host controls'],
                'correct': 1
            },
            {
                'question': 'What is recommended if a group experiences repeated disruptions?',
                'options': ['Close the meeting permanently', 'Hold a group conscience to discuss solutions', 'Change Zoom accounts', 'Assign every attendee as co-host'],
                'correct': 1
            },
            {
                'question': 'What message does the video end with?',
                'options': ['Stay safe and avoid Zoom.', 'Thank you for attending Back Porch training.', 'Happy hosting and keep coming back.', 'Goodbye and sign off.'],
                'correct': 2
            }
        ]
    }
}


@app.route("/quizzes")
@login_required
def quizzes_list():
    """Show available quizzes."""
    user = User.query.get(session["user_id"])
    
    # Get user's quiz attempts
    attempts = {}
    for quiz_id in QUIZZES.keys():
        attempt = QuizAttempt.query.filter_by(
            user_id=user.id,
            quiz_id=quiz_id,
            passed=True
        ).first()
        attempts[quiz_id] = attempt
    
    return render_template("quizzes_list.html", 
                         user=user, 
                         quizzes=QUIZZES,
                         attempts=attempts)


@app.route("/quiz/<quiz_id>")
@login_required
def quiz_view(quiz_id):
    """Show quiz video and questions."""
    if quiz_id not in QUIZZES:
        flash("Quiz not found.", "error")
        return redirect(url_for("quizzes_list"))
    
    user = User.query.get(session["user_id"])
    quiz = QUIZZES[quiz_id]
    
    # Check if user has already passed this quiz
    passed_attempt = QuizAttempt.query.filter_by(
        user_id=user.id,
        quiz_id=quiz_id,
        passed=True
    ).first()
    
    # Build video URL from config
    video_url = app.config['VIDEO_BASE_URL'] + quiz['video']
    
    return render_template("quiz_view.html", 
                         user=user, 
                         quiz=quiz,
                         video_url=video_url,
                         passed_attempt=passed_attempt)


@app.route("/quiz/<quiz_id>/submit", methods=["POST"])
@login_required
def quiz_submit(quiz_id):
    """Process quiz submission and award points."""
    if quiz_id not in QUIZZES:
        return jsonify({"error": "Quiz not found"}), 404
    
    user = User.query.get(session["user_id"])
    quiz = QUIZZES[quiz_id]
    
    # Get user answers from form
    user_answers = []
    for i in range(len(quiz['questions'])):
        answer = request.form.get(f'question_{i}')
        user_answers.append(int(answer) if answer else -1)
    
    # Calculate score
    correct_count = 0
    for i, question in enumerate(quiz['questions']):
        if i < len(user_answers) and user_answers[i] == question['correct']:
            correct_count += 1
    
    total_questions = len(quiz['questions'])
    score = int((correct_count / total_questions) * 100)
    passed = score >= 70
    
    # Award points if passed and not already completed
    points_awarded = 0
    if passed:
        previous_pass = QuizAttempt.query.filter_by(
            user_id=user.id,
            quiz_id=quiz_id,
            passed=True
        ).first()
        
        if not previous_pass:
            points_awarded = 50
            user.chair_points = (user.chair_points or 0) + points_awarded
    
    # Save attempt
    attempt = QuizAttempt(
        user_id=user.id,
        quiz_id=quiz_id,
        score=score,
        total_questions=total_questions,
        correct_answers=correct_count,
        passed=passed,
        answers=json.dumps(user_answers),
        points_awarded=points_awarded
    )
    db.session.add(attempt)
    db.session.commit()
    
    if passed:
        if points_awarded > 0:
            flash(f"Congratulations! You passed with {score}% and earned {points_awarded} ChairPoints! 🎉 Your certificate is ready to download!", "success")
        else:
            flash(f"Great job! You passed with {score}%. Your certificate is available for download.", "success")
    else:
        flash(f"You scored {score}%. You need 70% to pass. Please watch the video again and try again.", "warning")
    
    return redirect(url_for("quiz_view", quiz_id=quiz_id))


def generate_standard_certificate(user, attempt):
    """Generate a standard programmatic PDF certificate (fallback)"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from io import BytesIO
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Background border
    c.setStrokeColorRGB(0.06, 0.43, 0.46)  # Back Porch teal color
    c.setLineWidth(10)
    c.rect(30, 30, width - 60, height - 60, stroke=1, fill=0)
    
    # Inner decorative border
    c.setStrokeColorRGB(0.8, 0.8, 0.8)
    c.setLineWidth(2)
    c.rect(50, 50, width - 100, height - 100, stroke=1, fill=0)
    
    # Title
    c.setFont("Helvetica-Bold", 32)
    c.setFillColorRGB(0.06, 0.43, 0.46)
    c.drawCentredString(width / 2, height - 120, "CERTIFICATE OF COMPLETION")
    
    # Subtitle
    c.setFont("Helvetica", 16)
    c.setFillColorRGB(0, 0, 0)  # Pure black for readability
    c.drawCentredString(width / 2, height - 150, "Back Porch Online AA Meetings")
    
    # Decorative line
    c.setStrokeColorRGB(0.06, 0.43, 0.46)
    c.setLineWidth(2)
    c.line(150, height - 170, width - 150, height - 170)
    
    # "This certifies that"
    c.setFont("Helvetica", 14)
    c.setFillColorRGB(0, 0, 0)  # Pure black for readability
    c.drawCentredString(width / 2, height - 220, "This certifies that")
    
    # User name
    c.setFont("Helvetica-Bold", 28)
    c.setFillColorRGB(0.06, 0.43, 0.46)
    c.drawCentredString(width / 2, height - 260, user.display_name)
    
    # Achievement text
    c.setFont("Helvetica", 14)
    c.setFillColorRGB(0, 0, 0)  # Pure black for readability
    c.drawCentredString(width / 2, height - 300, "has successfully completed")
    
    # Training program title
    c.setFont("Helvetica-Bold", 20)
    c.setFillColorRGB(0.06, 0.43, 0.46)
    c.drawCentredString(width / 2, height - 335, "Back Porch Chairperson Training Program")
    
    # Both videos completed
    c.setFont("Helvetica", 13)
    c.setFillColorRGB(0, 0, 0)  # Pure black for readability
    c.drawCentredString(width / 2, height - 365, "Registration & Hosting Video Training")
    
    y_offset = 365
    
    # Score and details
    c.setFont("Helvetica", 12)
    c.setFillColorRGB(0, 0, 0)  # Pure black for readability
    c.drawCentredString(width / 2, height - y_offset - 40, 
                       f"Score: {attempt.score}% ({attempt.correct_answers}/{attempt.total_questions} correct)")
    c.drawCentredString(width / 2, height - y_offset - 60, 
                       f"ChairPoints Earned: {attempt.points_awarded}")
    
    # Date
    c.setFont("Helvetica-Oblique", 11)
    c.drawCentredString(width / 2, height - y_offset - 90, 
                       f"Completed on {attempt.completed_at.strftime('%B %d, %Y at %I:%M %p')}")
    
    # Bottom section - BP ID and signature line
    c.setFont("Helvetica", 10)
    c.setFillColorRGB(0, 0, 0)  # Pure black for readability
    
    # Left side - BP ID
    c.drawString(100, 120, f"Back Porch ID: {user.bp_id}")
    
    # Right side - Signature line
    c.line(width - 250, 135, width - 100, 135)
    c.drawString(width - 250, 120, "Authorized by Back Porch")
    
    # Footer
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColorRGB(0.2, 0.2, 0.2)  # Dark gray (more readable than light gray)
    c.drawCentredString(width / 2, 80, "The Real Back Porch - therealbackporch.com")
    c.drawCentredString(width / 2, 65, "Alcoholics Anonymous Online Meetings")
    
    # Certificate ID for verification
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.3, 0.3, 0.3)  # Medium gray for verification ID
    c.drawString(50, 50, f"Certificate ID: {attempt.id}-{user.id}-BP")
    
    c.save()
    buffer.seek(0)
    return buffer


@app.route("/certificate")
@app.route("/certificate/<int:user_id>")
@login_required
def quiz_certificate(user_id=None):
    """Generate a printable PDF certificate for a user - requires both quizzes to be completed.
    Admins can view any user's certificate. Regular users can only view their own."""
    
    # Determine which user's certificate to show
    if user_id is None:
        # No user_id provided, show current user's certificate
        user = User.query.get(session["user_id"])
    else:
        # user_id provided - check if current user is admin or viewing their own cert
        current_user_obj = User.query.get(session["user_id"])
        if not current_user_obj.is_admin and user_id != current_user_obj.id:
            flash("You can only view your own certificate.", "error")
            return redirect(url_for("dashboard"))
        
        user = User.query.get(user_id)
        if not user:
            flash("User not found.", "error")
            return redirect(url_for("admin_certificates") if current_user_obj.is_admin else url_for("dashboard"))
    
    # Check if user has completed BOTH quizzes
    all_quiz_ids = list(QUIZZES.keys())  # ['registration_quiz', 'hosting_quiz']
    passed_quizzes = QuizAttempt.query.filter_by(
        user_id=user.id,
        passed=True
    ).filter(QuizAttempt.quiz_id.in_(all_quiz_ids)).all()
    
    passed_quiz_ids = set([q.quiz_id for q in passed_quizzes])
    required_quiz_ids = set(all_quiz_ids)
    
    if not required_quiz_ids.issubset(passed_quiz_ids):
        missing_quizzes = required_quiz_ids - passed_quiz_ids
        missing_titles = [QUIZZES.get(qid, {}).get('title', qid) for qid in missing_quizzes]
        flash(f"Certificate not available. All video quizzes must be completed. Still needed: {', '.join(missing_titles)}", "warning")
        return redirect(url_for("quizzes_list"))
    
    # Get the most recent passed attempt for certificate details
    attempt = QuizAttempt.query.filter_by(
        user_id=user.id,
        passed=True
    ).order_by(QuizAttempt.completed_at.desc()).first()
    
    if not attempt:
        flash("No certificate data found.", "error")
        return redirect(url_for("quizzes_list"))
    
    # Try to use custom certificate image if available
    cert_image_path = os.path.join(app.root_path, 'static', 'img', 'certificate_template.png')
    
    if os.path.exists(cert_image_path):
        # Use custom certificate image with PIL
        try:
            from PIL import Image, ImageDraw, ImageFont
            
            # Open the certificate template
            img = Image.open(cert_image_path)
            draw = ImageDraw.Draw(img)
            
            # Get image dimensions
            img_width, img_height = img.size
            
            # ===== CERTIFICATE POSITIONING CONFIGURATION =====
            # Adjust these percentages to align text with your template's lines
            # Values are percentages of image height from top (0.0 = top, 1.0 = bottom)
            NAME_LINE_POSITION = 0.48   # Where the name line appears on your template
            DATE_LINE_POSITION = 0.78   # Where the date line appears on your template (moved down)
            PROGRAM_TEXT_POSITION = 0.72  # Additional program text position
            BP_ID_POSITION = 0.88        # BP ID position (moved down to avoid overlap)
            # =================================================
            
            # Try to load a nice font, fallback to default
            try:
                # Try different font sizes for name and date - SIGNIFICANTLY INCREASED FOR LEGIBILITY
                # Try bold fonts first for better visibility
                try:
                    name_font = ImageFont.truetype("arialbd.ttf", 96)  # Arial Bold, size 96
                    date_font = ImageFont.truetype("arialbd.ttf", 72)  # Arial Bold, size 72
                except:
                    # Fallback to regular Arial with larger sizes
                    name_font = ImageFont.truetype("arial.ttf", 96)
                    date_font = ImageFont.truetype("arial.ttf", 72)
            except:
                # Fallback to default font
                name_font = ImageFont.load_default()
                date_font = ImageFont.load_default()
            
            # Text color (dark, will work on light backgrounds)
            text_color = (0, 0, 0)  # Black - adjust if needed
            
            # NAME LINE - Align text baseline with template line
            name_text = user.display_name
            name_bbox = draw.textbbox((0, 0), name_text, font=name_font)
            name_width = name_bbox[2] - name_bbox[0]
            name_height = name_bbox[3] - name_bbox[1]
            name_x = (img_width - name_width) // 2
            # Position text so bottom of text sits ON the line at NAME_LINE_POSITION
            name_y = int(img_height * NAME_LINE_POSITION) - name_height
            draw.text((name_x, name_y), name_text, fill=text_color, font=name_font)
            
            # DATE LINE - Align text baseline with template line (positioned on Date line)
            date_text = attempt.completed_at.strftime('%B %d, %Y')
            date_bbox = draw.textbbox((0, 0), date_text, font=date_font)
            date_width = date_bbox[2] - date_bbox[0]
            date_height = date_bbox[3] - date_bbox[1]
            date_x = int(img_width * 0.18)  # Position at 18% from left edge (slightly left of center)
            # Position text so bottom of text sits ON the line at DATE_LINE_POSITION
            date_y = int(img_height * DATE_LINE_POSITION) - date_height
            draw.text((date_x, date_y), date_text, fill=text_color, font=date_font)
            
            # ADDITIONAL INFO - Not needed, template already has the program information
            
            # Convert PIL image to PDF
            img_buffer = BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Create PDF with the image
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter
            
            # Calculate scaling to fit letter size while maintaining aspect ratio
            aspect = img_width / img_height
            if aspect > (width / height):
                # Image is wider
                new_width = width
                new_height = width / aspect
            else:
                # Image is taller
                new_height = height
                new_width = height * aspect
            
            # Center the image
            x = (width - new_width) / 2
            y = (height - new_height) / 2
            
            # Draw the image
            if ImageReader is None:
                raise RuntimeError(
                    "Certificate image rendering is unavailable because Pillow/ImageReader could not be imported."
                )
            c.drawImage(ImageReader(img_buffer), x, y, width=new_width, height=new_height)
            
            c.save()
            buffer.seek(0)
            
        except Exception as e:
            app.logger.error(f"Error using custom certificate: {e}")
            # Fall back to generating standard certificate
            buffer = generate_standard_certificate(user, attempt)
    else:
        # No custom image, use standard certificate
        buffer = generate_standard_certificate(user, attempt)
    
    # Return PDF as downloadable file
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    filename = f'BP_Certificate_{user.display_name.replace(" ", "_")}.pdf'
    response.headers['Content-Disposition'] = f'inline; filename={filename}'
    
    return response


# ==========================
# EMAIL FUNCTIONS
# ==========================

def send_email(to, subject, body, ical_attachment=None, ical_filename=None):
    """Send an email using Flask-Mail with optional iCal attachment."""
    msg = Message(subject, recipients=[to], body=body)
    
    # Attach iCal file if provided
    if ical_attachment and ical_filename:
        msg.attach(
            ical_filename,
            "text/calendar",
            ical_attachment,
            headers=[('Content-Class', 'urn:content-classes:calendarmessage')]
        )
    
    try:
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Email send failed: {e}")
        return False


def generate_meeting_ical(meeting, chair_name=None):
    """Generate iCal data for a meeting."""
    cal = Calendar()
    cal.add('prodid', '-//Back Porch Meetings//Chairperson Scheduler//EN')
    cal.add('version', '2.0')
    cal.add('method', 'REQUEST')
    
    event = Event()
    event.add('summary', f"CHAIR: {meeting.title}")
    event.add('description', f"""You are scheduled to chair this meeting.

Meeting: {meeting.title}
Type: {getattr(meeting, 'meeting_type', 'Regular')}
Description: {meeting.description or 'Standard meeting format'}
Zoom Link: {meeting.zoom_link or 'Contact admin'}

Please join 10-15 minutes early to set up.""")
    
    # Set start and end times
    start_datetime = datetime.combine(meeting.event_date, meeting.start_time or datetime.min.time())
    
    # Calculate end time
    if meeting.end_time:
        end_datetime = datetime.combine(meeting.event_date, meeting.end_time)
    else:
        # Default to 1 hour duration if no end time specified
        end_datetime = start_datetime + timedelta(hours=1)
    
    event.add('dtstart', start_datetime)
    event.add('dtend', end_datetime)
    event.add('location', meeting.zoom_link or 'Online')
    
    if chair_name:
        event.add('organizer', chair_name)
    
    # Add reminder alarms
    # 24 hours before
    alarm_24h = Alarm()
    alarm_24h.add('action', 'DISPLAY')
    alarm_24h.add('description', f'Reminder: You are chairing {meeting.title} tomorrow')
    alarm_24h.add('trigger', timedelta(hours=-24))
    event.add_component(alarm_24h)
    
    # 1 hour before
    alarm_1h = Alarm()
    alarm_1h.add('action', 'DISPLAY')
    alarm_1h.add('description', f'Starting soon: {meeting.title} in 1 hour')
    alarm_1h.add('trigger', timedelta(hours=-1))
    event.add_component(alarm_1h)
    
    cal.add_component(event)
    return cal.to_ical()

def send_chair_reminder(meeting_id, hours_before=24):
    """Send reminder email to chair before meeting."""
    meeting = Meeting.query.get(meeting_id)
    if not meeting or not meeting.chair_signup:
        return False
    
    chair = meeting.chair_signup.user
    
    # Determine subject and content based on timing
    if hours_before == 24:
        subject = f"Reminder: You're chairing tomorrow — {meeting.title}"
        timing_text = "tomorrow"
        prep_text = """
Preparation reminders:
• Review the meeting format and agenda
• Test your Zoom connection and screen sharing
• Prepare any announcements or readings
• Have backup contact information ready

Resources:
• Chairperson Guidelines: [Available in your profile]
• Zoom Host Guide: [Available in your profile]
"""
    else:  # 1 hour before
        subject = f"Starting soon: {meeting.title} in 1 hour"
        timing_text = "in about 1 hour"
        prep_text = """
Final checklist:
• Join Zoom 10-15 minutes early
• Enable waiting room if needed
• Have your opening/closing ready
• Check that screen sharing works

"""
    
    body = f"""Hi {chair.display_name},

Your meeting starts {timing_text}:

📅 {meeting.title}
📍 {meeting.event_date.strftime('%A, %B %d, %Y')}
🕐 {meeting.start_time.strftime('%I:%M %p')}
📝 {meeting.description or 'Standard format'}
📧 Meeting Type: {getattr(meeting, 'meeting_type', 'Regular')}

{prep_text}
🔗 Zoom Link: {meeting.zoom_link or 'Contact admin for meeting link'}

Thank you for serving the Back Porch community!

— Back Porch Meetings System
"""
    
    # Generate iCal attachment
    ical_data = generate_meeting_ical(meeting, chair.display_name)
    ical_filename = f"meeting_{meeting.id}_reminder.ics"
    
    return send_email(chair.email, subject, body, ical_data, ical_filename)


def send_meeting_confirmations():
    """Send confirmation emails for newly assigned meetings."""
    from datetime import datetime, timedelta
    
    # Find meetings in the next 30 days that were recently assigned (created in last 24 hours)
    cutoff_time = datetime.utcnow() - timedelta(hours=24)
    upcoming_start = datetime.utcnow().date()
    upcoming_end = (datetime.utcnow() + timedelta(days=30)).date()
    
    recent_signups = (
        ChairSignup.query
        .join(Meeting)
        .filter(
            ChairSignup.created_at >= cutoff_time,
            Meeting.event_date >= upcoming_start,
            Meeting.event_date <= upcoming_end
        )
        .all()
    )
    
    sent_count = 0
    for signup in recent_signups:
        if send_chair_confirmation(signup):
            sent_count += 1
    
    return sent_count


def send_chair_confirmation(chair_signup):
    """Send confirmation email when someone signs up to chair."""
    meeting = chair_signup.meeting
    chair = chair_signup.user
    
    subject = f"Confirmed: You're chairing {meeting.title}"
    body = f"""Hi {chair.display_name},

Great news! You're confirmed to chair this meeting:

📅 {meeting.title}
📍 {meeting.event_date.strftime('%A, %B %d, %Y')}
🕐 {meeting.start_time.strftime('%I:%M %p')}
📝 {meeting.description or 'Standard meeting format'}
📧 Meeting Type: {getattr(meeting, 'meeting_type', 'Regular')}

What happens next:
• You'll get a reminder email 24 hours before
• Another reminder 1 hour before the meeting
• Meeting details and Zoom link: {meeting.zoom_link or 'Will be provided'}

📎 An iCal event is attached to this email - add it to your calendar!

Need help? Check out the chairperson resources in your profile or contact the admin team.

Thank you for stepping up to serve!

— Back Porch Meetings System
"""
    
    # Generate iCal attachment
    ical_data = generate_meeting_ical(meeting, chair.display_name)
    ical_filename = f"chairperson_meeting_{meeting.event_date.strftime('%Y%m%d')}_{meeting.id}.ics"
    
    return send_email(chair.email, subject, body, ical_data, ical_filename)


def check_and_send_reminders():
    """Check for meetings that need reminder emails and send them."""
    from datetime import datetime, timedelta
    
    now = datetime.utcnow()
    
    # Find meetings needing 24-hour reminders
    tomorrow_start = now + timedelta(hours=23)
    tomorrow_end = now + timedelta(hours=25)
    
    meetings_24h = (
        Meeting.query
        .join(ChairSignup)
        .filter(
            Meeting.event_date == tomorrow_start.date(),
            Meeting.start_time >= tomorrow_start.time(),
            Meeting.start_time <= tomorrow_end.time()
        )
        .all()
    )
    
    # Find meetings needing 1-hour reminders
    hour_start = now + timedelta(minutes=55)
    hour_end = now + timedelta(minutes=65)
    
    meetings_1h = (
        Meeting.query
        .join(ChairSignup)
        .filter(
            Meeting.event_date == now.date(),
            Meeting.start_time >= hour_start.time(),
            Meeting.start_time <= hour_end.time()
        )
        .all()
    )
    
    sent_count = 0
    
    # Send 24-hour reminders
    for meeting in meetings_24h:
        if send_chair_reminder(meeting.id, hours_before=24):
            sent_count += 1
            print(f"Sent 24h reminder for meeting {meeting.id}")
    
    # Send 1-hour reminders
    for meeting in meetings_1h:
        if send_chair_reminder(meeting.id, hours_before=1):
            sent_count += 1
            print(f"Sent 1h reminder for meeting {meeting.id}")
    
    return sent_count
    
    send_email(chair.email, subject, body)

def send_open_slot_reminder():
    """Send weekly reminder about open chair slots."""
    tomorrow = date.today() + timedelta(days=1)
    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= tomorrow, Meeting.event_date <= tomorrow + timedelta(days=7))
        .filter(Meeting.is_open == True)
        .filter(Meeting.chair_signup == None)
        .all()
    )
    
    if not meetings:
        return
    
    # Send to all users (or admins could send manually)
    users = User.query.filter_by(is_admin=False).all()
    subject = "Back Porch: Open Chair Positions This Week"
    body = """Dear Back Porch Chairperson,

There are open chair positions available this week. Please visit the chairperson portal to sign up:

"""
    for m in meetings:
        body += f"- {m.title} on {m.event_date.strftime('%A, %B %d')} at {m.start_time.strftime('%I:%M %p')}\n"
    
    body += "\nThank you for your service!\n\nBack Porch Meetings"
    
    for user in users:
        send_email(user.email, subject, body)

def send_day_of_chair_reminders(run_date: date = None):
    """Send reminder emails to chairs on the morning of their scheduled meeting day.
    If run_date is provided, use it; otherwise, use today's date.
    """
    target_day = run_date or date.today()
    meetings = (
        Meeting.query
        .filter(Meeting.event_date == target_day)
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )

    sent = 0
    for m in meetings:
        if not m.chair_signup or not m.chair_signup.user:
            continue
        chair = m.chair_signup.user
        subject = f"Heads up! You're hosting today — {m.title}"
        body = f"""Hi {chair.display_name},

    Heads up — you're scheduled to host today! Here are the details:

    • Meeting: {m.title}
    • Date: {m.event_date.strftime('%A, %B %d, %Y')}
    • Time: {m.start_time.strftime('%I:%M %p')}
    • Description: {m.description or 'N/A'}
    • Zoom Link: {m.zoom_link or 'Contact group for link'}

    Thanks for stepping up to serve the Back Porch community!

    — Back Porch Meetings
    """
        if send_email(chair.email, subject, body):
            sent += 1

    return sent


def send_availability_confirmation_email(user, volunteer_date, time_preference):
    """Send confirmation email when user volunteers for a date."""
    time_text = {
        'any': 'any time',
        'morning': 'morning (before 12 PM)',
        'afternoon': 'afternoon (12 PM - 6 PM)', 
        'evening': 'evening (after 6 PM)'
    }.get(time_preference, 'any time')
    
    subject = f"Chairperson Volunteer Confirmation — {volunteer_date.strftime('%B %d, %Y')}"
    body = f"""Hi {user.display_name},

Thank you for volunteering to chair a Back Porch meeting!

Your volunteer details:
• Date: {volunteer_date.strftime('%A, %B %d, %Y')}
• Time Preference: {time_text}

What happens next:
- We'll let you know if a meeting gets scheduled for this date
- You'll receive a reminder email 24 hours before any meeting
- If plans change, you can contact the admin team

Thanks for your willingness to serve the Back Porch community!

— Back Porch Meetings
"""
    
    send_email(user.email, subject, body)


def award_chair_points_for_completed_meetings():
    """Award 1 ChairPoint to users who chaired meetings that have passed.
    Runs daily to award points for meetings from the previous day.
    """
    yesterday = date.today() - timedelta(days=1)
    
    # Find all meetings from yesterday that had a chair
    completed_meetings = (
        Meeting.query
        .filter(Meeting.event_date == yesterday)
        .join(ChairSignup)
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )
    
    points_awarded = 0
    for meeting in completed_meetings:
        if meeting.chair_signup and meeting.chair_signup.user:
            chair = meeting.chair_signup.user
            # Award 1 point
            chair.chair_points = (chair.chair_points or 0) + 1
            points_awarded += 1
            app.logger.info(f"Awarded 1 ChairPoint to {chair.display_name} (BP-{1000 + chair.id}) for meeting {meeting.title}")
    
    if points_awarded > 0:
        db.session.commit()
        app.logger.info(f"Total ChairPoints awarded: {points_awarded}")
    
    return points_awarded


# ==========================
# ADMIN ROUTES
# ==========================

@app.route("/admin/certificates")
@admin_required
def admin_certificates():
    """Admin page to view all certificate recipients"""
    # Get all passed quiz attempts with user info
    certificate_recipients = db.session.query(
        User.id,
        User.display_name,
        User.email,
        db.func.count(QuizAttempt.id).label('quizzes_passed'),
        db.func.max(QuizAttempt.completed_at).label('latest_completion')
    ).join(
        QuizAttempt, User.id == QuizAttempt.user_id
    ).filter(
        QuizAttempt.passed == True
    ).group_by(
        User.id
    ).having(
        db.func.count(QuizAttempt.id) >= 2  # Both quizzes passed
    ).order_by(
        db.desc('latest_completion')
    ).all()
    
    # Get detailed quiz attempts for each recipient
    recipients_with_details = []
    for recipient in certificate_recipients:
        user_attempts = QuizAttempt.query.filter_by(
            user_id=recipient.id,
            passed=True
        ).order_by(QuizAttempt.completed_at.desc()).all()
        
        # Get the user object to access bp_id property
        user_obj = User.query.get(recipient.id)
        
        recipients_with_details.append({
            'user_id': recipient.id,
            'display_name': recipient.display_name,
            'bp_id': user_obj.bp_id if user_obj else 'N/A',
            'email': recipient.email,
            'quizzes_passed': recipient.quizzes_passed,
            'latest_completion': recipient.latest_completion,
            'attempts': user_attempts
        })
    
    return render_template(
        'admin_certificates.html',
        recipients=recipients_with_details,
        total_recipients=len(recipients_with_details)
    )

@app.route("/admin/meetings")
@admin_required
def admin_meetings():
    # Get search and filter parameters
    search_query = request.args.get('search', '').strip()
    meeting_type_filter = request.args.get('meeting_type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    chair_status = request.args.get('chair_status', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)  # Configurable page size
    
    # Validate per_page limits for performance
    per_page = min(max(per_page, 10), 100)  # Between 10 and 100
    
    # Start with base query with optimized loading
    query = Meeting.query.options(
        db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user)
    )
    
    # Apply search filter with index-friendly queries
    if search_query:
        query = query.filter(
            db.or_(
                Meeting.title.ilike(f'%{search_query}%'),
                Meeting.description.ilike(f'%{search_query}%')
            )
        )
    
    # Apply meeting type filter (uses index)
    if meeting_type_filter:
        query = query.filter(Meeting.meeting_type == meeting_type_filter)
    
    # Apply date range filters (uses indexes)
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Meeting.event_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Meeting.event_date <= to_date)
        except ValueError:
            pass
    
    # Apply chair status filter
    if chair_status == 'has_chair':
        query = query.join(ChairSignup)
    elif chair_status == 'needs_chair':
        query = query.filter(~Meeting.chair_signup.has())
    
    # Order by date and time with index optimization
    query = query.order_by(Meeting.event_date.desc(), Meeting.start_time.desc())
    
    # Paginate results for performance
    pagination = query.paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    meetings = pagination.items
    
    # Get meeting type options for filter dropdown (cached)
    meeting_types = cache.get('meeting_types_list')
    if meeting_types is None:
        meeting_types = db.session.query(Meeting.meeting_type).distinct().all()
        meeting_types = [mt[0] for mt in meeting_types if mt[0]]
        cache.set('meeting_types_list', meeting_types, timeout=3600)  # Cache for 1 hour
    
    # Get all users for chair assignment dropdown
    all_users = User.query.order_by(User.display_name.asc()).all()
    
    return render_template(
        "admin_meetings.html", 
        meetings=meetings,
        pagination=pagination,
        ics_source=SOURCE_MEETINGS_ICS_URL, 
        web_source=SOURCE_MEETINGS_WEB_URL, 
        static_schedule_enabled=STATIC_SCHEDULE_ENABLED,
        meeting_types=meeting_types,
        all_users=all_users,
        current_filters={
            'search': search_query,
            'meeting_type': meeting_type_filter,
            'date_from': date_from,
            'date_to': date_to,
            'chair_status': chair_status,
            'page': page,
            'per_page': per_page
        }
    )


@app.route("/admin/import-ics")
@admin_required
def admin_import_ics():
    """Admin endpoint to import meetings from configured ICS URL."""
    url = SOURCE_MEETINGS_ICS_URL
    if not url:
        flash("ICS source URL is not configured.", "warning")
        return redirect(url_for('admin_meetings'))
    try:
        count = import_meetings_from_ics(url, replace_future=True)
        flash(f"Imported {count} meetings from ICS.", "success")
    except Exception as e:
        flash(f"ICS import failed: {e}", "danger")
    return redirect(url_for('admin_meetings'))


@app.route("/admin/import-web")
@admin_required
def admin_import_web():
    """Admin endpoint to import meetings by scraping the configured web page."""
    url = SOURCE_MEETINGS_WEB_URL
    if not url:
        flash("Web source URL is not configured.", "warning")
        return redirect(url_for('admin_meetings'))
    try:
        count = import_meetings_from_webpage(url, weeks=int(request.args.get("weeks", 12)), replace_future=True)
        flash(f"Imported {count} meetings from website.", "success")
    except Exception as e:
        flash(f"Website import failed: {e}", "danger")
    return redirect(url_for('admin_meetings'))


@app.route("/admin/seed-static")
@admin_required
def admin_seed_static():
    """Admin endpoint to seed meetings from built-in static schedule."""
    if not STATIC_SCHEDULE_ENABLED:
        flash("Static schedule seeding is disabled.", "warning")
        return redirect(url_for('admin_meetings'))
    try:
        count = seed_meetings_from_static_schedule(weeks=int(request.args.get("weeks", 12)), replace_future=True)
        flash(f"Created {count} meetings from static schedule.", "success")
    except Exception as e:
        flash(f"Static schedule seed failed: {e}", "danger")
    return redirect(url_for('admin_meetings'))


@app.route("/admin/diagnostics")
@admin_required
def admin_diagnostics():
    """Quick health diagnostics for DB and mail configuration."""
    engine_url = str(db.engine.url)
    inspector = db.inspect(db.engine)
    tables = inspector.get_table_names()
    counts = {}
    for t in tables:
        try:
            counts[t] = db.session.execute(db.text(f"SELECT COUNT(*) AS c FROM `{t}`")).scalar()
        except Exception:
            counts[t] = None

    mail_settings = {
        "MAIL_SERVER": app.config.get("MAIL_SERVER"),
        "MAIL_PORT": app.config.get("MAIL_PORT"),
        "MAIL_USE_TLS": app.config.get("MAIL_USE_TLS"),
        "MAIL_USE_SSL": app.config.get("MAIL_USE_SSL"),
        "MAIL_DEFAULT_SENDER": app.config.get("MAIL_DEFAULT_SENDER"),
        "MAIL_USERNAME_set": bool(app.config.get("MAIL_USERNAME")),
        "MAIL_PASSWORD_set": bool(app.config.get("MAIL_PASSWORD")),
    }

    return render_template(
        "admin_diagnostics.html",
        engine_url=engine_url,
        tables=tables,
        counts=counts,
        mail_settings=mail_settings,
        ics_source=SOURCE_MEETINGS_ICS_URL,
        web_source=SOURCE_MEETINGS_WEB_URL,
    )


@app.route("/admin/reports/all-users")
@admin_required
def admin_all_users_report():
    """Complete report of all registered users with chair points and meeting counts.
    Shows users who earned points from quizzes even if they haven't chaired yet.
    Optional CSV export with ?format=csv
    """
    fmt = request.args.get("format", "html").lower()
    
    # Get all users with their chair points and meeting counts
    all_users = db.session.query(
        User.id,
        User.display_name,
        User.email,
        User.chair_points,
        User.profile_image,
        User.created_at,
        func.count(ChairSignup.id).label('total_meetings')
    ).outerjoin(ChairSignup).group_by(User.id).order_by(User.chair_points.desc(), User.display_name).all()
    
    # Calculate totals
    total_users = len(all_users)
    total_points = sum(user.chair_points for user in all_users)
    total_meetings = sum(user.total_meetings for user in all_users)
    users_with_points = sum(1 for user in all_users if user.chair_points > 0)
    users_with_meetings = sum(1 for user in all_users if user.total_meetings > 0)
    
    if fmt == "csv":
        import csv
        from io import StringIO
        si = StringIO()
        writer = csv.DictWriter(si, fieldnames=[
            "bp_id", "display_name", "email", "chair_points", "total_meetings", "joined_date"
        ])
        writer.writeheader()
        for user in all_users:
            writer.writerow({
                "bp_id": f"BP-{1000 + user.id}",
                "display_name": user.display_name,
                "email": user.email,
                "chair_points": user.chair_points,
                "total_meetings": user.total_meetings,
                "joined_date": user.created_at.strftime('%Y-%m-%d') if user.created_at else ''
            })
        resp = make_response(si.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = 'attachment; filename=all_users_report.csv'
        return resp
    
    # HTML view
    return render_template("admin_all_users_report.html", 
                         all_users=all_users,
                         total_users=total_users,
                         total_points=total_points,
                         total_meetings=total_meetings,
                         users_with_points=users_with_points,
                         users_with_meetings=users_with_meetings)


@app.route("/admin/reports/chairpoints-leaderboard")
@admin_required
def admin_chairpoints_leaderboard():
    """Full ChairPoints Leaderboard Report - All chairpersons sorted by points.
    Optional CSV export with ?format=csv
    """
    fmt = request.args.get("format", "html").lower()
    
    # Get all users with chair points, sorted by points descending
    leaderboard = db.session.query(
        User.id,
        User.display_name,
        User.chair_points,
        User.profile_image,
        User.email,
        func.count(ChairSignup.id).label('total_meetings')
    ).outerjoin(ChairSignup).filter(
        User.chair_points > 0
    ).group_by(User.id).order_by(User.chair_points.desc()).all()
    
    if fmt == "csv":
        import csv
        from io import StringIO
        si = StringIO()
        writer = csv.DictWriter(si, fieldnames=[
            "rank", "bp_id", "display_name", "email", "chair_points", "total_meetings"
        ])
        writer.writeheader()
        for idx, chair in enumerate(leaderboard, start=1):
            writer.writerow({
                "rank": idx,
                "bp_id": f"BP-{1000 + chair.id}",
                "display_name": chair.display_name,
                "email": chair.email,
                "chair_points": chair.chair_points,
                "total_meetings": chair.total_meetings
            })
        resp = make_response(si.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = 'attachment; filename=chairpoints_leaderboard.csv'
        return resp
    
    # HTML view
    return render_template("admin_chairpoints_leaderboard.html", leaderboard=leaderboard)


@app.route("/admin/reports/chair-schedule")
@admin_required
def admin_report_chair_schedule():
    """Report: Chair Schedule by date/time with registered name and BP-ID.
    Optional CSV export with ?format=csv
    """
    fmt = request.args.get("format", "html").lower()
    meetings = (
        Meeting.query
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )

    # Build rows
    rows = []
    for m in meetings:
        chair_name = m.chair_signup.display_name_snapshot if m.chair_signup else None
        bp_id = m.chair_signup.user.bp_id if (m.chair_signup and m.chair_signup.user) else None
        rows.append({
            "date": m.event_date.strftime('%Y-%m-%d'),
            "weekday": m.event_date.strftime('%A'),
            "start": m.start_time.strftime('%H:%M'),
            "title": m.title,
            "chair_name": chair_name,
            "bp_id": bp_id,
            "is_open": m.is_open and not m.has_chair,
        })

    if fmt == "csv":
        import csv
        from io import StringIO
        si = StringIO()
        writer = csv.DictWriter(si, fieldnames=[
            "date","weekday","start","title","chair_name","bp_id","is_open"
        ])
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
        resp = make_response(si.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = 'attachment; filename=chair_schedule.csv'
        return resp

    # Group by date for HTML
    grouped = {}
    for r in rows:
        grouped.setdefault(r["date"], []).append(r)

    return render_template("admin_chair_schedule.html", grouped=grouped)


@app.route("/admin/reports/monthly-pdf")
@admin_required
def admin_monthly_pdf():
    """Generate a monthly PDF report listing meetings chronologically with date, time, chair name, and title."""
    today = get_eastern_today()
    year = int(request.args.get("year", today.year))
    month = int(request.args.get("month", today.month))

    # Get all meetings in month
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)
    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= start_date, Meeting.event_date <= end_date)
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )

    # Build PDF in memory
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter

    title = f"Back Porch Monthly Report - {calendar.month_name[month]} {year}"
    c.setFont("Helvetica-Bold", 14)
    c.drawString(1*inch, height - 1*inch, title)
    c.setFont("Helvetica", 10)
    c.drawString(1*inch, height - 1.25*inch, f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} ")

    y = height - 1.6*inch
    line_height = 0.22 * inch

    # Table header
    c.setFont("Helvetica-Bold", 11)
    c.drawString(1*inch, y, "Date")
    c.drawString(2.5*inch, y, "Time")
    c.drawString(4*inch, y, "Chair")
    c.drawString(6*inch, y, "Meeting")
    y -= line_height
    c.setFont("Helvetica", 10)

    if not meetings:
        c.drawString(1*inch, y, "No meetings scheduled for this month.")
    else:
        for m in meetings:
            # Check page break
            if y < 1*inch:
                c.showPage()
                y = height - 1*inch
                c.setFont("Helvetica-Bold", 11)
                c.drawString(1*inch, y, "Date")
                c.drawString(2.5*inch, y, "Time")
                c.drawString(4*inch, y, "Chair")
                c.drawString(6*inch, y, "Meeting")
                y -= line_height
                c.setFont("Helvetica", 10)

            date_str = m.event_date.strftime('%Y-%m-%d (%a)')
            time_str = m.start_time.strftime('%I:%M %p') + (f" - {m.end_time.strftime('%I:%M %p')}" if m.end_time else "")
            chair_str = (
                f"{m.chair_signup.display_name_snapshot} ({m.chair_signup.user.bp_id})"
                if m.chair_signup and m.chair_signup.user else "No chair yet"
            )
            title_str = m.title

            c.drawString(1*inch, y, date_str)
            c.drawString(2.5*inch, y, time_str)
            c.drawString(4*inch, y, chair_str[:28])  # prevent overflow
            c.drawString(6*inch, y, title_str[:28])
            y -= line_height

    c.showPage()
    c.save()
    pdf = buf.getvalue()
    buf.close()

    resp = make_response(pdf)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = (
        f"attachment; filename=monthly_report_{year}_{month}.pdf"
    )
    return resp


@app.route("/admin/reports/monthly")
@admin_required
def admin_monthly_html():
    """Admin user statistics report showing chair assignments and completion rates."""
    today = get_eastern_today()
    
    # Get all registered users (non-admin)
    users = User.query.filter_by(is_admin=False).order_by(User.display_name).all()
    
    user_stats = []
    for user in users:
        # Count past meetings chaired (to date)
        past_chaired = (
            db.session.query(Meeting)
            .join(ChairSignup, Meeting.id == ChairSignup.meeting_id)
            .filter(
                ChairSignup.user_id == user.id,
                Meeting.event_date < today
            )
            .count()
        )
        
        # Count future meetings chaired
        future_chaired = (
            db.session.query(Meeting)
            .join(ChairSignup, Meeting.id == ChairSignup.meeting_id)
            .filter(
                ChairSignup.user_id == user.id,
                Meeting.event_date >= today
            )
            .count()
        )
        
        # Get future meeting details
        future_meetings = (
            db.session.query(Meeting)
            .join(ChairSignup, Meeting.id == ChairSignup.meeting_id)
            .filter(
                ChairSignup.user_id == user.id,
                Meeting.event_date >= today
            )
            .order_by(Meeting.event_date.asc())
            .all()
        )
        
        # Calculate percentage (past + future)
        total_chaired = past_chaired + future_chaired
        
        # Only include users who have chaired or will chair
        if total_chaired > 0:
            user_stats.append({
                'user': user,
                'past_chaired': past_chaired,
                'future_chaired': future_chaired,
                'total_chaired': total_chaired,
                'future_meetings': future_meetings
            })
    
    # Calculate totals
    total_past = sum(s['past_chaired'] for s in user_stats)
    total_future = sum(s['future_chaired'] for s in user_stats)
    total_all = total_past + total_future
    total_chairpoints = sum(s['user'].chair_points or 0 for s in user_stats)
    
    # Count unfilled future meetings
    unfilled_future = (
        Meeting.query
        .filter(
            Meeting.event_date >= today,
            ~Meeting.chair_signup.has()
        )
        .count()
    )
    
    return render_template(
        "admin_monthly.html",
        user_stats=user_stats,
        total_past=total_past,
        total_future=total_future,
        total_all=total_all,
        total_chairpoints=total_chairpoints,
        unfilled_future=unfilled_future,
        today=today,
    )


@app.route("/admin/meetings/new", methods=["GET", "POST"])
@admin_required
def admin_meeting_new():
    form = MeetingForm()
    if form.validate_on_submit():
        meeting = Meeting(
            title=form.title.data.strip(),
            description=form.description.data.strip() if form.description.data else None,
            zoom_link=form.zoom_link.data.strip() if form.zoom_link.data else None,
            event_date=form.event_date.data,
            start_time=form.start_time.data,
            end_time=form.end_time.data,
            is_open=form.is_open.data,
            gender_restriction=form.gender_restriction.data or None,
            meeting_type=form.meeting_type.data,
        )
        db.session.add(meeting)
        db.session.commit()
        flash("Meeting created.", "success")
        return redirect(url_for("admin_meetings"))
    return render_template("admin_meeting_form.html", form=form, mode="new")


@app.route("/admin/meetings/<int:meeting_id>/edit", methods=["GET", "POST"])
@admin_required
def admin_meeting_edit(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    form = MeetingForm(obj=meeting)

    if form.validate_on_submit():
        meeting.title = form.title.data.strip()
        meeting.description = form.description.data.strip() if form.description.data else None
        meeting.zoom_link = form.zoom_link.data.strip() if form.zoom_link.data else None
        meeting.event_date = form.event_date.data
        meeting.start_time = form.start_time.data
        meeting.end_time = form.end_time.data
        meeting.is_open = form.is_open.data
        meeting.gender_restriction = form.gender_restriction.data or None
        meeting.meeting_type = form.meeting_type.data
        db.session.commit()
        flash("Meeting updated.", "success")
        return redirect(url_for("admin_meetings"))

    return render_template("admin_meeting_form.html", form=form, mode="edit", meeting=meeting)


@app.route("/admin/meetings/<int:meeting_id>/delete", methods=["POST"])
@admin_required
def admin_meeting_delete(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    db.session.delete(meeting)
    db.session.commit()
    flash("Meeting deleted.", "info")
    return redirect(url_for("admin_meetings"))


@app.route("/admin/meetings/<int:meeting_id>/clear-chair", methods=["POST"])
@admin_required
def admin_meeting_clear_chair(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    if meeting.chair_signup:
        db.session.delete(meeting.chair_signup)
        db.session.commit()
        flash("Chair signup cleared.", "info")
    else:
        flash("This meeting currently has no chair to clear.", "warning")
    return redirect(url_for("admin_meetings"))


@app.route("/admin/reminders/send", methods=["POST"])
@admin_required
def admin_send_reminders():
    """Manually trigger reminder emails for upcoming meetings."""
    try:
        sent_count = check_and_send_reminders()
        flash(f"Sent {sent_count} reminder emails.", "success")
    except Exception as e:
        flash(f"Error sending reminders: {str(e)}", "danger")
    
    return redirect(url_for("admin_meetings"))


@app.route("/admin/reminders/test/<int:meeting_id>", methods=["POST"])
@admin_required
def admin_test_reminder(meeting_id):
    """Send a test reminder email for a specific meeting."""
    meeting = Meeting.query.get_or_404(meeting_id)
    
    if not meeting.chair_signup:
        flash("Cannot send reminder - no chair assigned to this meeting.", "warning")
        return redirect(url_for("admin_meetings"))
    
    try:
        if send_chair_reminder(meeting_id, hours_before=24):
            flash(f"Test reminder sent to {meeting.chair_signup.user.display_name}.", "success")
        else:
            flash("Failed to send test reminder.", "danger")
    except Exception as e:
        flash(f"Error sending test reminder: {str(e)}", "danger")
    
    return redirect(url_for("admin_meetings"))


@app.route("/admin/confirmations/send", methods=["POST"])
@admin_required
def admin_send_confirmations():
    """Manually send confirmation emails for recent signups."""
    try:
        sent_count = send_meeting_confirmations()
        flash(f"Sent {sent_count} confirmation emails.", "success")
    except Exception as e:
        flash(f"Error sending confirmations: {str(e)}", "danger")
    
    return redirect(url_for("admin_meetings"))


@app.route("/offline")
def offline():
    """Offline fallback page for PWA."""
    return render_template("offline.html")


# ==========================
# API: Interactive Calendar
# ==========================

@app.route("/api/day-meetings")
def api_day_meetings():
    """Return JSON list of meetings for a given date (YYYY-MM-DD)."""
    date_str = request.args.get("date")
    try:
        target = datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"error": "Invalid date"}), 400

    meetings = (
        Meeting.query
        .filter(Meeting.event_date == target)
        .order_by(Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )
    user = get_current_user()
    data = []
    for m in meetings:
        eligible = True
        if m.gender_restriction in ('male','female'):
            eligible = bool(user and user.gender == m.gender_restriction)
        data.append({
            "id": m.id,
            "title": m.title,
            "time": m.start_time.strftime('%I:%M %p'),
            "has_chair": bool(m.chair_signup),
            "chair_name": m.chair_signup.user.display_name if (m.chair_signup and m.chair_signup.user) else None,
            "chair_profile_url": url_for('profile_image', user_id=m.chair_signup.user_id) if (m.chair_signup and m.chair_signup.user and m.chair_signup.user.profile_image) else None,
            "is_open": m.is_open and not m.chair_signup,
            "eligible": eligible,
            "detail_url": url_for('meeting_detail', meeting_id=m.id),
        })
    return jsonify({"date": date_str, "meetings": data})


@app.route("/api/week-meetings")
def api_week_meetings():
    """Return JSON list of meetings for a date range."""
    start_str = request.args.get("start")
    end_str = request.args.get("end")
    try:
        start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"error": "Invalid date"}), 400

    meetings = (
        Meeting.query
        .filter(Meeting.event_date >= start_date, Meeting.event_date <= end_date)
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .all()
    )
    
    data = []
    for m in meetings:
        data.append({
            "id": m.id,
            "title": m.title,
            "date": m.event_date.strftime('%m/%d'),
            "time": m.start_time.strftime('%I:%M %p') if m.start_time else '',
            "chair_name": m.chair_signup.user.display_name if (m.chair_signup and m.chair_signup.user) else None,
            "chair_profile_url": url_for('profile_image', user_id=m.chair_signup.user_id) if (m.chair_signup and m.chair_signup.user and m.chair_signup.user.profile_image) else None,
        })
    return jsonify({"meetings": data})


@app.route("/api/meetings/<int:meeting_id>/claim", methods=["POST"])
def api_meeting_claim(meeting_id):
    """Claim chairing for an open meeting. Requires login."""
    user = get_current_user()
    if not user:
        return jsonify({"error": "login_required"}), 401

    meeting = Meeting.query.options(db.joinedload(Meeting.chair_signup)).get_or_404(meeting_id)
    if not meeting.is_open:
        return jsonify({"error": "not_open"}), 400
    if meeting.chair_signup:
        return jsonify({"error": "already_has_chair"}), 400

    # Enforce gender restriction if set
    if meeting.gender_restriction in ('male','female'):
        if not user.gender:
            return jsonify({"error": "gender_required"}), 400
        if user.gender != meeting.gender_restriction:
            return jsonify({"error": "not_eligible"}), 403

    signup = ChairSignup(
        meeting_id=meeting.id,
        user_id=user.id,
        display_name_snapshot=user.display_name,
        notes=None,
    )
    db.session.add(signup)
    db.session.commit()

    return jsonify({
        "ok": True,
        "meeting_id": meeting.id,
        "chair_name": user.display_name,
    })


@app.route("/api/volunteer-date", methods=["POST"])
def api_volunteer_date():
    """API endpoint for volunteering to chair on a specific date."""
    user = get_current_user()
    if not user:
        return jsonify({"error": "login_required"}), 401

    data = request.get_json()
    if not data or 'date' not in data:
        return jsonify({"error": "date_required"}), 400

    try:
        volunteer_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"error": "invalid_date_format"}), 400

    # Check if date is in the past
    if volunteer_date < date.today():
        return jsonify({"error": "past_date"}), 400

    # Check if user already volunteered for this date
    existing = ChairpersonAvailability.query.filter_by(
        user_id=user.id, 
        volunteer_date=volunteer_date, 
        is_active=True
    ).first()
    
    if existing:
        return jsonify({"error": "already_volunteered"}), 400

    # Check if there are already meetings scheduled for this date
    existing_meetings = Meeting.query.filter_by(event_date=volunteer_date).count()
    
    # Create availability record
    availability = ChairpersonAvailability(
        user_id=user.id,
        volunteer_date=volunteer_date,
        time_preference=data.get('time_preference', 'any'),
        notes=data.get('notes', '').strip() if data.get('notes') else None,
        display_name_snapshot=user.display_name
    )
    
    db.session.add(availability)
    db.session.commit()

    # Send confirmation email
    try:
        send_availability_confirmation_email(user, volunteer_date, availability.time_preference)
    except Exception as e:
        app.logger.error(f"Failed to send availability confirmation email: {e}")

    return jsonify({
        "ok": True,
        "message": "Thank you for volunteering! You'll receive an email confirmation.",
        "date": volunteer_date.strftime('%Y-%m-%d'),
        "existing_meetings": existing_meetings
    })


@app.route("/volunteer-date/<date_str>", methods=["GET", "POST"])
def volunteer_date_page(date_str):
    """Page for volunteering to chair on a specific date."""
    try:
        volunteer_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash("Invalid date format.", "danger")
        return redirect(url_for('calendar_view'))

    if volunteer_date < date.today():
        flash("Cannot volunteer for past dates.", "warning")
        return redirect(url_for('calendar_view'))

    user = get_current_user()
    if not user:
        flash("Please log in to volunteer for chairperson duties.", "warning")
        return redirect(url_for("login", next=request.path))

    # Check for existing meetings on this date
    meetings = Meeting.query.filter_by(event_date=volunteer_date).all()
    
    # Check if user already volunteered
    existing_availability = ChairpersonAvailability.query.filter_by(
        user_id=user.id, 
        volunteer_date=volunteer_date, 
        is_active=True
    ).first()

    form = ChairpersonAvailabilityForm()
    
    if form.validate_on_submit():
        if existing_availability:
            flash("You have already volunteered for this date.", "info")
        else:
            availability = ChairpersonAvailability(
                user_id=user.id,
                volunteer_date=volunteer_date,
                time_preference=form.time_preference.data,
                notes=form.notes.data.strip() if form.notes.data else None,
                display_name_snapshot=user.display_name
            )
            
            db.session.add(availability)
            db.session.commit()

            # Send confirmation email
            try:
                send_availability_confirmation_email(user, volunteer_date, availability.time_preference)
            except Exception as e:
                app.logger.error(f"Failed to send availability confirmation email: {e}")

            flash("Thank you for volunteering to chair! You'll receive an email confirmation.", "success")
            return redirect(url_for('calendar_view'))

    return render_template('volunteer_date.html', 
                         date=volunteer_date, 
                         meetings=meetings, 
                         existing_availability=existing_availability,
                         form=form)


@app.route("/volunteer/bulk-template")
@login_required
def volunteer_bulk_template():
    """Download a CSV template for bulk volunteer date registration."""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header row
    writer.writerow(['Date (YYYY-MM-DD)', 'Time Preference', 'Notes (Optional)'])
    
    # Write example rows with next 7 days
    today = date.today()
    for i in range(1, 8):
        example_date = today + timedelta(days=i)
        time_pref = ['morning', 'afternoon', 'evening', 'any'][i % 4]
        writer.writerow([example_date.strftime('%Y-%m-%d'), time_pref, ''])
    
    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=volunteer_dates_template.csv'
    
    return response


@app.route("/volunteer/bulk-template-excel")
@login_required
def volunteer_bulk_template_excel():
    """Download an Excel template for bulk volunteer date registration."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        flash("Excel template not available. Please use CSV format.", "warning")
        return redirect(url_for('volunteer_bulk_template'))
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volunteer Dates"
    
    # Style for headers
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Write headers
    headers = ['Date (YYYY-MM-DD)', 'Time Preference', 'Notes (Optional)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    
    # Write example rows with next 7 days
    today = date.today()
    for i in range(1, 8):
        example_date = today + timedelta(days=i)
        time_pref = ['morning', 'afternoon', 'evening', 'any'][i % 4]
        
        ws.cell(row=i+1, column=1, value=example_date.strftime('%Y-%m-%d'))
        ws.cell(row=i+1, column=2, value=time_pref)
        ws.cell(row=i+1, column=3, value='')
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=volunteer_dates_template.xlsx'
    
    return response


@app.route("/volunteer/bulk-upload", methods=["GET", "POST"])
@login_required
def volunteer_bulk_upload():
    """Bulk upload volunteer dates via CSV or Excel file."""
    user = get_current_user()
    
    if request.method == "GET":
        return render_template('volunteer_bulk_upload.html')
    
    # POST: Process file upload
    if 'csv_file' not in request.files:
        flash("No file uploaded.", "danger")
        return redirect(url_for('volunteer_bulk_upload'))
    
    file = request.files['csv_file']
    
    if file.filename == '':
        flash("No file selected.", "danger")
        return redirect(url_for('volunteer_bulk_upload'))
    
    # Check file extension
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        flash("Please upload a CSV or Excel (.xlsx) file.", "danger")
        return redirect(url_for('volunteer_bulk_upload'))
    
    try:
        # Parse file based on type
        rows = []
        
        if filename.endswith('.csv'):
            # Parse CSV
            import csv
            from io import TextIOWrapper
            
            stream = TextIOWrapper(file.stream, encoding='utf-8')
            csv_reader = csv.DictReader(stream)
            
            # Validate headers
            expected_headers = {'Date (YYYY-MM-DD)', 'Time Preference', 'Notes (Optional)'}
            if not expected_headers.issubset(set(csv_reader.fieldnames or [])):
                flash("Invalid CSV format. Please download the template and use the correct headers.", "danger")
                return redirect(url_for('volunteer_bulk_upload'))
            
            rows = list(csv_reader)
            
        else:
            # Parse Excel
            try:
                import openpyxl
            except ImportError:
                flash("Excel file support not available. Please use CSV format.", "danger")
                return redirect(url_for('volunteer_bulk_upload'))
            
            workbook = openpyxl.load_workbook(file.stream)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            expected_headers = ['Date (YYYY-MM-DD)', 'Time Preference', 'Notes (Optional)']
            
            if headers[:3] != expected_headers:
                flash("Invalid Excel format. Please download the template and use the correct headers.", "danger")
                return redirect(url_for('volunteer_bulk_upload'))
            
            # Convert rows to dict format
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Skip empty rows
                    rows.append({
                        'Date (YYYY-MM-DD)': str(row[0]) if row[0] else '',
                        'Time Preference': str(row[1]) if row[1] else '',
                        'Notes (Optional)': str(row[2]) if row[2] else ''
                    })
        
        # Process rows
        success_count = 0
        error_count = 0
        errors = []
        today = date.today()
        
        for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (header is row 1)
            try:
                # Parse date
                date_str = row['Date (YYYY-MM-DD)'].strip()
                if not date_str:
                    continue  # Skip empty rows
                
                volunteer_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                # Validate date is not in the past
                if volunteer_date < today:
                    errors.append(f"Row {row_num}: Date {date_str} is in the past")
                    error_count += 1
                    continue
                
                # Validate time preference
                time_pref = row['Time Preference'].strip().lower()
                valid_prefs = ['any', 'morning', 'afternoon', 'evening']
                if time_pref not in valid_prefs:
                    errors.append(f"Row {row_num}: Invalid time preference '{time_pref}'. Must be one of: {', '.join(valid_prefs)}")
                    error_count += 1
                    continue
                
                # Check for existing volunteer record
                existing = ChairpersonAvailability.query.filter_by(
                    user_id=user.id,
                    volunteer_date=volunteer_date,
                    is_active=True
                ).first()
                
                if existing:
                    errors.append(f"Row {row_num}: Already volunteered for {date_str}")
                    error_count += 1
                    continue
                
                # Create availability record
                notes = row['Notes (Optional)'].strip() if row.get('Notes (Optional)') else None
                
                availability = ChairpersonAvailability(
                    user_id=user.id,
                    volunteer_date=volunteer_date,
                    time_preference=time_pref,
                    notes=notes[:500] if notes else None,  # Limit to 500 chars
                    display_name_snapshot=user.display_name
                )
                
                db.session.add(availability)
                success_count += 1
                
                # Send confirmation email for each date
                try:
                    send_availability_confirmation_email(user, volunteer_date, time_pref)
                except Exception as e:
                    app.logger.error(f"Failed to send bulk upload confirmation email: {e}")
                
            except ValueError as e:
                errors.append(f"Row {row_num}: Invalid date format '{date_str}'. Use YYYY-MM-DD")
                error_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
        
        # Commit all successful records
        if success_count > 0:
            db.session.commit()
            flash(f"Successfully registered {success_count} volunteer date(s)!", "success")
        
        # Show errors if any
        if error_count > 0:
            error_msg = f"{error_count} row(s) had errors:<br>" + "<br>".join(errors[:10])
            if len(errors) > 10:
                error_msg += f"<br>...and {len(errors) - 10} more errors"
            flash(error_msg, "warning")
        
        if success_count == 0 and error_count == 0:
            flash("No valid dates found in CSV file.", "info")
        
        return redirect(url_for('profile'))
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error processing CSV file: {str(e)}", "danger")
        app.logger.error(f"Bulk upload error: {e}")
        return redirect(url_for('volunteer_bulk_upload'))


# ==========================
# CLI: init DB + default admin
# ==========================
# ANALYTICS ROUTES
# ==========================

@app.route("/admin/analytics")
@require_login
def admin_analytics():
    """Advanced analytics dashboard for administrators."""
    # Get filter parameter (default to 30 days)
    days_filter = request.args.get('days', '30')
    try:
        days = int(days_filter)
        if days not in [30, 60, 90]:
            days = 30
    except ValueError:
        days = 30
    
    # Summary statistics
    total_meetings = Meeting.query.count()
    active_chairpersons = User.query.join(ChairSignup).distinct().count()
    covered_meetings = Meeting.query.join(ChairSignup).count()
    coverage_percentage = round((covered_meetings / total_meetings * 100) if total_meetings > 0 else 0)
    
    # Average signup time (days before meeting)
    signups_with_days = db.session.query(
        func.avg(
            func.datediff(
                func.concat(Meeting.event_date, ' ', func.coalesce(Meeting.start_time, '00:00:00')),
                ChairSignup.created_at
            )
        )
    ).join(Meeting).scalar()
    avg_signup_time = f"{int(signups_with_days or 0)} days" if signups_with_days else "N/A"
    
    # Attendance trends (last 12 weeks)
    end_date = date.today()
    start_date = end_date - timedelta(weeks=12)
    
    attendance_trends = []
    labels = []
    covered_data = []
    total_data = []
    
    current_date = start_date
    while current_date <= end_date:
        week_end = current_date + timedelta(days=6)
        
        week_meetings = Meeting.query.filter(
            Meeting.event_date >= current_date,
            Meeting.event_date <= week_end
        ).count()
        
        week_covered = Meeting.query.filter(
            Meeting.event_date >= current_date,
            Meeting.event_date <= week_end
        ).join(ChairSignup).count()
        
        labels.append(current_date.strftime('%m/%d'))
        total_data.append(week_meetings)
        covered_data.append(week_covered)
        
        current_date += timedelta(weeks=1)
    
    attendance_trends_data = {
        'labels': labels,
        'total': total_data,
        'covered': covered_data
    }
    
    # Meeting types distribution
    meeting_types = db.session.query(
        Meeting.meeting_type,
        func.count(Meeting.id)
    ).group_by(Meeting.meeting_type).all()
    
    meeting_types_data = {
        'labels': [mt[0] or 'Regular' for mt in meeting_types],
        'data': [mt[1] for mt in meeting_types]
    }
    
    # Popular time slots
    time_slots = db.session.query(
        func.hour(Meeting.start_time).label('hour'),
        func.count(Meeting.id).label('count')
    ).filter(Meeting.start_time.isnot(None)).group_by('hour').order_by('hour').all()
    
    # ChairPoints Leaderboard - Top 20 chairs by points
    leaderboard = db.session.query(
        User.id,
        User.display_name,
        User.chair_points,
        User.profile_image
    ).filter(User.chair_points > 0).order_by(User.chair_points.desc()).limit(20).all()
    
    time_slots_data = {
        'labels': [f"{ts[0]}:00" for ts in time_slots],
        'data': [ts[1] for ts in time_slots]
    }
    
    # Weekly distribution (day of week)
    weekly_dist = db.session.query(
        func.dayofweek(Meeting.event_date).label('day'),
        func.count(Meeting.id).label('count')
    ).group_by('day').order_by('day').all()
    
    day_names = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    weekly_data = [0] * 7
    for day, count in weekly_dist:
        weekly_data[day-1] = count  # MySQL dayofweek returns 1-7
    
    weekly_distribution_data = {
        'labels': day_names,
        'data': weekly_data
    }
    
    # Top chairpersons with filtered period
    filter_start_date = date.today() - timedelta(days=days)
    top_chairpersons = db.session.query(
        User.display_name,
        func.count(ChairSignup.id).label('total_meetings')
    ).join(ChairSignup).join(Meeting).filter(
        Meeting.event_date >= filter_start_date
    ).group_by(User.id, User.display_name).order_by(func.count(ChairSignup.id).desc()).limit(15).all()
    
    # Convert to list of dicts for easier template access
    top_chairpersons_list = [
        {
            'display_name': chair[0],
            'total_meetings': chair[1],
            'period_count': chair[1]  # Same as total in this filtered query
        }
        for chair in top_chairpersons
    ]
    
    # Meetings needing chairs (next 30 days)
    uncovered_meetings = Meeting.query.filter(
        Meeting.event_date >= date.today(),
        Meeting.event_date <= date.today() + timedelta(days=30),
        ~Meeting.chair_signup.has()
    ).order_by(Meeting.event_date.asc(), Meeting.start_time.asc()).limit(10).all()
    
    # ChairPoints Leaderboard - Top 20 chairs by points
    leaderboard = db.session.query(
        User.id,
        User.display_name,
        User.chair_points,
        User.profile_image
    ).filter(User.chair_points > 0).order_by(User.chair_points.desc()).limit(20).all()
    
    return render_template(
        "admin_analytics.html",
        total_meetings=total_meetings,
        active_chairpersons=active_chairpersons,
        coverage_percentage=coverage_percentage,
        avg_signup_time=avg_signup_time,
        attendance_trends_data=json.dumps(attendance_trends_data),
        meeting_types_data=json.dumps(meeting_types_data),
        time_slots_data=json.dumps(time_slots_data),
        weekly_distribution_data=json.dumps(weekly_distribution_data),
        top_chairpersons=top_chairpersons_list,
        uncovered_meetings=uncovered_meetings,
        days_filter=days,
        leaderboard=leaderboard
    )


@app.route("/admin/security")
@require_login
def admin_security():
    """Security dashboard for administrators."""
    current_user = get_current_user()
    if not current_user.is_admin:
        abort(403)
    
    # Security overview stats
    total_users = User.query.count()
    active_sessions = AuditLog.query.filter(
        AuditLog.action == 'login_success',
        AuditLog.created_at > datetime.now(timezone.utc) - timedelta(hours=24)
    ).count()
    
    failed_logins_today = AuditLog.query.filter(
        AuditLog.action.like('login_failure%'),
        AuditLog.created_at > datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    ).count()
    
    recent_backups = BackupLog.query.filter(
        BackupLog.created_at > datetime.now(timezone.utc) - timedelta(weeks=1)
    ).count()
    
    # Recent audit logs (last 100)
    audit_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(100).all()
    
    # All users with role information
    users = User.query.order_by(User.created_at.desc()).all()
    
    # Recent backup logs
    backup_logs = BackupLog.query.order_by(BackupLog.created_at.desc()).limit(20).all()
    
    return render_template(
        "admin_security.html",
        total_users=total_users,
        active_sessions=active_sessions,
        failed_logins_today=failed_logins_today,
        recent_backups=recent_backups,
        audit_logs=audit_logs,
        users=users,
        backup_logs=backup_logs
    )


@app.route("/admin/security/backup", methods=["POST"])
@require_login
def admin_create_backup():
    """Create a manual database backup."""
    current_user = get_current_user()
    if not current_user.is_admin:
        abort(403)
    
    try:
        # Check rate limiting
        if not check_rate_limit('manual_backup', current_user.id, 5):  # Max 5 per hour
            return jsonify({"success": False, "error": "Too many backup requests. Please try again later."}), 429
        
        backup_log = backup_database('manual', current_user.id)
        return jsonify({"success": True, "backup_id": backup_log.id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/admin/security/users/<int:user_id>/unlock", methods=["POST"])
@require_login
def admin_unlock_user(user_id):
    """Unlock a user account."""
    current_user = get_current_user()
    if not current_user.is_admin:
        abort(403)
    
    try:
        user = User.query.get_or_404(user_id)
        user.unlock_account()
        
        log_audit_event('admin_unlock_account', current_user.id, 'user', user_id, {
            'target_user': user.email,
            'unlocked_by': current_user.email
        })
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/admin/meetings/bulk-delete", methods=["POST"])
@require_login
def admin_bulk_delete_meetings():
    """Bulk delete meetings."""
    current_user = get_current_user()
    if not current_user.is_admin:
        abort(403)
    
    try:
        data = request.get_json()
        meeting_ids = data.get('meeting_ids', [])
        
        if not meeting_ids:
            return jsonify({"success": False, "error": "No meetings selected"}), 400
        
        # Validate all meeting IDs exist
        meetings = Meeting.query.filter(Meeting.id.in_(meeting_ids)).all()
        if len(meetings) != len(meeting_ids):
            return jsonify({"success": False, "error": "Some meetings not found"}), 404
        
        # Check rate limiting
        if not check_rate_limit('bulk_delete', current_user.id, 3):  # Max 3 bulk deletes per hour
            return jsonify({"success": False, "error": "Too many bulk operations. Please try again later."}), 429
        
        # Delete meetings and associated chair signups
        deleted_count = 0
        for meeting in meetings:
            # Log the deletion
            log_audit_event('meeting_delete', current_user.id, 'meeting', meeting.id, {
                'title': meeting.title,
                'date': meeting.event_date.isoformat(),
                'bulk_operation': True
            })
            
            db.session.delete(meeting)
            deleted_count += 1
        
        db.session.commit()
        
        # Invalidate caches
        invalidate_meeting_caches()
        
        log_audit_event('bulk_delete_meetings', current_user.id, details={
            'meeting_count': deleted_count,
            'meeting_ids': meeting_ids
        })
        
        return jsonify({"success": True, "deleted_count": deleted_count})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/admin/meetings/export")
@require_login
def admin_export_meetings():
    """Export selected meetings as CSV or ICS."""
    current_user = get_current_user()
    if not current_user.is_admin:
        abort(403)
    
    meeting_ids = request.args.getlist('ids[]')
    format_type = request.args.get('format', 'csv')
    
    if not meeting_ids:
        flash("No meetings selected for export.", "warning")
        return redirect(url_for('admin_meetings'))
    
    # Get meetings with chair information
    meetings = (
        Meeting.query
        .filter(Meeting.id.in_(meeting_ids))
        .options(db.joinedload(Meeting.chair_signup).joinedload(ChairSignup.user))
        .order_by(Meeting.event_date.asc(), Meeting.start_time.asc())
        .all()
    )
    
    if format_type == 'ics':
        return export_meetings_ics(meetings)
    else:
        return export_meetings_csv(meetings)


def export_meetings_csv(meetings):
    """Export meetings as CSV."""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'ID', 'Title', 'Description', 'Date', 'Start Time', 'End Time', 
        'Meeting Type', 'Chair Name', 'Chair Email', 'Chair BP ID', 'Zoom Link'
    ])
    
    # Write meeting data
    for meeting in meetings:
        chair = meeting.chair_signup
        writer.writerow([
            meeting.id,
            meeting.title,
            meeting.description or '',
            meeting.event_date.strftime('%Y-%m-%d'),
            meeting.start_time.strftime('%H:%M') if meeting.start_time else '',
            meeting.end_time.strftime('%H:%M') if meeting.end_time else '',
            meeting.meeting_type,
            chair.display_name_snapshot if chair else '',
            chair.user.email if chair else '',
            chair.user.bp_id if chair else '',
            meeting.zoom_link or ''
        ])
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=meetings_export_{date.today().isoformat()}.csv'
    
    log_audit_event('export_meetings', get_current_user().id, details={
        'format': 'csv',
        'meeting_count': len(meetings)
    })
    
    return response


def export_meetings_ics(meetings):
    """Export meetings as ICS calendar file."""
    cal = Calendar()
    cal.add('prodid', '-//Back Porch Meetings Export//backporchmeetings.org//')
    cal.add('version', '2.0')
    cal.add('x-wr-calname', 'Back Porch Meetings Export')
    
    for meeting in meetings:
        event = Event()
        start_dt = datetime.combine(meeting.event_date, meeting.start_time)
        end_dt = datetime.combine(meeting.event_date, meeting.end_time) if meeting.end_time else start_dt + timedelta(hours=1)
        
        event.add('summary', meeting.title)
        event.add('dtstart', start_dt)
        event.add('dtend', end_dt)
        event.add('location', meeting.zoom_link or 'Online')
        
        description = meeting.description or ''
        if meeting.chair_signup:
            description += f"\n\nChair: {meeting.chair_signup.display_name_snapshot} ({meeting.chair_signup.user.bp_id})"
        
        event.add('description', description)
        event.add('uid', f"export-meeting-{meeting.id}@backporchmeetings.org")
        
        cal.add_component(event)
    
    response = make_response(cal.to_ical())
    response.headers['Content-Type'] = 'text/calendar; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=meetings_export_{date.today().isoformat()}.ics'
    
    log_audit_event('export_meetings', get_current_user().id, details={
        'format': 'ics',
        'meeting_count': len(meetings)
    })
    
    return response


# Admin: Bulk Email to Chairs
@app.route("/admin/chairs/email", methods=["GET", "POST"])
@admin_required
def admin_email_chairs():
    """Send bulk email to all active chairpersons."""
    if request.method == "POST":
        subject = request.form.get("subject", "").strip()
        message = request.form.get("message", "").strip()
        recipient_filter = request.form.get("recipient_filter", "all")
        
        if not subject or not message:
            flash("Subject and message are required.", "danger")
            return redirect(url_for("admin_email_chairs"))
        
        # Get recipients based on filter
        if recipient_filter == "all":
            # All users who have ever chaired
            users = User.query.join(ChairSignup).distinct().all()
        elif recipient_filter == "active":
            # Users with upcoming meetings
            users = (
                User.query
                .join(ChairSignup)
                .join(Meeting)
                .filter(Meeting.event_date >= date.today())
                .distinct()
                .all()
            )
        elif recipient_filter == "past30":
            # Users who chaired in last 30 days
            cutoff = date.today() - timedelta(days=30)
            users = (
                User.query
                .join(ChairSignup)
                .join(Meeting)
                .filter(Meeting.event_date >= cutoff, Meeting.event_date < date.today())
                .distinct()
                .all()
            )
        else:
            flash("Invalid recipient filter.", "danger")
            return redirect(url_for("admin_email_chairs"))
        
        # Send emails
        sent_count = 0
        failed_count = 0
        for user in users:
            try:
                send_email(user.email, subject, message)
                sent_count += 1
            except Exception as e:
                failed_count += 1
                app.logger.error(f"Failed to send email to {user.email}: {e}")
        
        flash(f"Sent {sent_count} emails. {failed_count} failed.", "success" if failed_count == 0 else "warning")
        log_audit_event('bulk_email_chairs', get_current_user().id, details={
            'recipient_filter': recipient_filter,
            'sent_count': sent_count,
            'failed_count': failed_count
        })
        return redirect(url_for("admin_meetings"))
    
    # GET request - show form
    all_count = User.query.join(ChairSignup).distinct().count()
    active_count = User.query.join(ChairSignup).join(Meeting).filter(Meeting.event_date >= date.today()).distinct().count()
    cutoff = date.today() - timedelta(days=30)
    past30_count = User.query.join(ChairSignup).join(Meeting).filter(Meeting.event_date >= cutoff, Meeting.event_date < date.today()).distinct().count()
    
    return render_template("admin_email_chairs.html", 
                         all_count=all_count,
                         active_count=active_count,
                         past30_count=past30_count)


# Admin: Manual Chair Assignment
@app.route("/admin/meetings/<int:meeting_id>/assign-chair", methods=["POST"])
@admin_required
def admin_assign_chair(meeting_id):
    """Manually assign a chair to a meeting."""
    meeting = Meeting.query.get_or_404(meeting_id)
    assignment_type = request.form.get("assignment_type", "existing")
    notes = request.form.get("notes", "").strip()
    
    user = None
    chair_name = None
    chair_email = None
    
    if assignment_type == "existing":
        # Existing user assignment
        user_id = request.form.get("user_id", type=int)
        if not user_id:
            flash("Please select a chairperson.", "danger")
            return redirect(url_for("admin_meetings"))
        
        user = User.query.get_or_404(user_id)
        chair_name = user.display_name
        chair_email = user.email
        
    else:
        # New chair by email
        chair_name = request.form.get("chair_name", "").strip()
        chair_email = request.form.get("chair_email", "").strip()
        
        if not chair_name or not chair_email:
            flash("Please provide both name and email for the new chair.", "danger")
            return redirect(url_for("admin_meetings"))
        
        # Check if user already exists with this email
        user = User.query.filter_by(email=chair_email).first()
        if user:
            # User exists, use their account
            chair_name = user.display_name
        # If user doesn't exist, we'll create a signup with just the email
    
    # Remove existing chair if any
    if meeting.chair_signup:
        db.session.delete(meeting.chair_signup)
    
    # Create new signup
    signup = ChairSignup(
        meeting_id=meeting.id,
        user_id=user.id if user else None,
        display_name_snapshot=chair_name,
        notes=notes or "Manually assigned by admin"
    )
    db.session.add(signup)
    db.session.commit()
    
    flash(f"Assigned {chair_name} to chair this meeting.", "success")
    log_audit_event('assign_chair', get_current_user().id, details={
        'meeting_id': meeting_id,
        'chair_name': chair_name,
        'chair_email': chair_email
    })
    
    # Send confirmation email
    try:
        subject = f"Chairperson Assignment: {meeting.title}"
        body = f"""Hello {chair_name},

You have been assigned to chair the following meeting:

Meeting: {meeting.title}
Date: {meeting.event_date.strftime('%A, %B %d, %Y')}
Time: {meeting.start_time.strftime('%I:%M %p') if meeting.start_time else 'TBD'}

{f'Zoom Link: {meeting.zoom_link}' if meeting.zoom_link else ''}

Notes: {notes or 'None'}

If you have any questions, please contact the administrator.

Thank you for your service!
"""
        send_email(chair_email, subject, body)
    except Exception as e:
        app.logger.error(f"Failed to send assignment email: {e}")
    
    return redirect(url_for("admin_meetings"))


# Admin: Export Chair Activity Report (CSV)
@app.route("/admin/reports/chair-activity")
@admin_required
def admin_chair_activity_report():
    """Export detailed chair activity report as CSV."""
    days = request.args.get('days', 90, type=int)
    cutoff = date.today() - timedelta(days=days)
    
    # Get all users who have chaired
    chair_data = db.session.query(
        User.id,
        User.display_name,
        User.email,
        User.bp_id,
        User.phone,
        db.func.count(ChairSignup.id).label('total_meetings'),
        db.func.count(db.case(
            (Meeting.event_date >= cutoff, 1)
        )).label('recent_meetings'),
        db.func.min(Meeting.event_date).label('first_meeting'),
        db.func.max(Meeting.event_date).label('last_meeting')
    ).join(ChairSignup).join(Meeting).group_by(User.id).order_by(db.func.count(ChairSignup.id).desc()).all()
    
    # Create CSV
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'Chairperson Name', 'Email', 'BP ID', 'Phone', 
        'Total Meetings', f'Last {days} Days', 'First Meeting', 'Last Meeting'
    ])
    
    for row in chair_data:
        writer.writerow([
            row.display_name,
            row.email,
            row.bp_id or '',
            row.phone or '',
            row.total_meetings,
            row.recent_meetings,
            row.first_meeting.strftime('%Y-%m-%d') if row.first_meeting else '',
            row.last_meeting.strftime('%Y-%m-%d') if row.last_meeting else ''
        ])
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=chair_activity_{days}days_{date.today().isoformat()}.csv'
    
    log_audit_event('export_chair_activity', get_current_user().id, details={'days': days})
    
    return response


# ==========================

@app.cli.command("init-db")
def init_db_command():
    """
    Initialize the database and create a default admin user.
    Run with: flask --app app.py init-db
    """
    db.create_all()

    if not User.query.filter_by(email="chair.admin@backporchmeetings.org").first():
        admin = User(
            display_name="Back Porch Admin",
            email="chair.admin@backporchmeetings.org",
            is_admin=True,
            agreed_guidelines=True
        )
        admin.set_password("changeme123")
        db.session.add(admin)
        db.session.commit()
        print("Created default admin: chair.admin@backporchmeetings.org / changeme123")
    else:
        print("Admin already exists.")

    # Schedule weekly open slots reminder (only in development)
    # Skip scheduler in Heroku release phase
    if scheduler and not os.environ.get('DYNO'):
        scheduler.add_job(
            send_open_slot_reminder,
            CronTrigger(day_of_week='sun', hour=10),
            id='weekly-open-slots',
            replace_existing=True
        )
        print("Scheduled weekly open slots reminder.")
        
        # Schedule daily ChairPoints award at 1 AM
        scheduler.add_job(
            award_chair_points_for_completed_meetings,
            CronTrigger(hour=1, minute=0),
            id='daily-chairpoints-award',
            replace_existing=True
        )
        print("Scheduled daily ChairPoints award job.")

    print("Database initialized.")


@app.cli.command("upgrade-schema")
def upgrade_schema_command():
    """
    Upgrade DB schema to add gender columns if they don't exist.
    Intended for MySQL; uses ALTER TABLE guarded by information_schema checks.
    Run with: flask --app app.py upgrade-schema
    """
    engine = db.engine
    conn = engine.connect()
    inspector = db.inspect(engine)

    # Add users.gender if missing
    if 'users' in inspector.get_table_names():
        cols = [c['name'] if isinstance(c, dict) else c['name'] for c in inspector.get_columns('users')]
        if 'gender' not in cols:
            try:
                conn.execute(db.text("ALTER TABLE users ADD COLUMN gender VARCHAR(10) NULL"))
                print("Added users.gender")
            except Exception as e:
                print(f"Skipped adding users.gender: {e}")

    # Add meetings.gender_restriction if missing
    if 'meetings' in inspector.get_table_names():
        cols = [c['name'] if isinstance(c, dict) else c['name'] for c in inspector.get_columns('meetings')]
        if 'gender_restriction' not in cols:
            try:
                conn.execute(db.text("ALTER TABLE meetings ADD COLUMN gender_restriction VARCHAR(10) NULL"))
                print("Added meetings.gender_restriction")
            except Exception as e:
                print(f"Skipped adding meetings.gender_restriction: {e}")

    conn.close()
    print("Schema upgrade complete.")


@app.cli.command("import-ics")
def import_ics_command():
    """Import meetings from ICS URL defined in env var SOURCE_MEETINGS_ICS_URL.
    Run with: flask --app app.py import-ics
    """
    url = SOURCE_MEETINGS_ICS_URL
    if not url:
        print("SOURCE_MEETINGS_ICS_URL is not set.")
        return
    try:
        count = import_meetings_from_ics(url, replace_future=True)
        print(f"Imported {count} meetings from ICS.")
    except Exception as e:
        print(f"Import failed: {e}")


@app.cli.command("seed-schedule")
def seed_schedule_command():
    """Seed meetings from the built-in static schedule for the next N weeks.
    Usage: flask --app app.py seed-schedule
    Optionally set WEEKS env variable to control horizon (default 12).
    """
    weeks = int(os.environ.get("WEEKS", "12"))
    count = seed_meetings_from_static_schedule(weeks=weeks, replace_future=True)
    print(f"Seeded {count} meetings from static schedule.")


@app.cli.command("import-web")
def import_web_command():
    """Import meetings by scraping the website URL defined in env var SOURCE_MEETINGS_WEB_URL.
    Usage: flask --app app.py import-web
    Optionally set WEEKS env variable to control horizon (default 12).
    """
    url = SOURCE_MEETINGS_WEB_URL
    if not url:
        print("SOURCE_MEETINGS_WEB_URL is not set.")
        return
    weeks = int(os.environ.get("WEEKS", "12"))
    try:
        count = import_meetings_from_webpage(url, weeks=weeks, replace_future=True)
        print(f"Imported {count} meetings from website.")
    except Exception as e:
        print(f"Website import failed: {e}")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True)

